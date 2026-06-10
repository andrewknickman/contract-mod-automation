import os
import argparse
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from file_selection import (
    choose_existing_file, choose_save_file, read_excel_auto, EXCEL_FILETYPES,
)


# ─── Runtime-selected inputs ──────────────────────────────────────────────────
# Assigned in main() from CLI arguments or file picker dialogs.
j17_file = None
j1_previous_file_input = None
j1_current_file = None
eis_billing_file = None
j17_updated_file = None

# ─── Configuration ────────────────────────────────────────────────────────────
current_option_period = 5
DEFAULT_CURRENT_MONTH = "December"
current_month = DEFAULT_CURRENT_MONTH


def extract_month_from_filename(filename):
    match = re.search(r'EIS Billing Detail - (\w+) (\d+)', filename.replace('_', ' '))
    if match:
        month_name = match.group(1)
        year = match.group(2)[-2:]
        return f"{year}{month_name.upper()}"
    return ""


def extract_month_name_from_filename(filename):
    match = re.search(r'EIS Billing Detail - (\w+) (\d+)', filename.replace('_', ' '))
    return match.group(1).title() if match else None


def get_catalog_sheet_name(option_period):
    return f"2B_Opt Pd {option_period} Catalog"


def load_existing_j17_data(filepath):
    catalog_df = read_excel_auto(filepath, sheet_name='Catalog of Subscription CLINs')
    active_df = read_excel_auto(filepath, sheet_name='J.17 Active Subscriptions')
    expired_df = read_excel_auto(filepath, sheet_name='J.17 Expired Subscriptions')
    return catalog_df, active_df, expired_df


def load_draft_j1_catalog(filepath, option_period):
    sheet_name = get_catalog_sheet_name(option_period)
    df = read_excel_auto(filepath, sheet_name=sheet_name)
    return df


def load_billing_data(filepath):
    bi_df = read_excel_auto(filepath, sheet_name='Billing Invoice (BI) Detail')
    ba_df = read_excel_auto(filepath, sheet_name='Billing Adjustment (BA) Detail')
    return bi_df, ba_df


def filter_subscription_rows(df):
    keywords = ['Year', '12 Month', 'Annual', 'License', 'Subscription', 'Duration']
    pattern = '|'.join(keywords)
    mask = df['Verizon Case Description'].astype(str).str.contains(pattern, case=False, na=False)
    return df[mask].copy()


def get_new_catalog_rows(existing_catalog_df, draft_j1_df):
    subscription_rows = filter_subscription_rows(draft_j1_df)
    
    if len(existing_catalog_df) > 0:
        existing_case_numbers = set(existing_catalog_df['Case Number'].dropna().astype(str).tolist())
        new_rows = subscription_rows[~subscription_rows['Case Number'].astype(str).isin(existing_case_numbers)]
    else:
        new_rows = subscription_rows
    
    return new_rows.copy()


def vlookup_bi_with_catalog(bi_df, catalog_df):
    bi_df = bi_df.copy()
    bi_df['individual case basis code number'] = bi_df['individual case basis code number'].astype(str)
    catalog_df = catalog_df.copy()
    catalog_df['Case Number'] = catalog_df['Case Number'].astype(str)
    
    merged = bi_df.merge(
        catalog_df[['Case Number', 'EIS CLIN', 'EIS CLIN Name', 'Verizon Case Description', 'SRE Pricing Element']],
        left_on='individual case basis code number',
        right_on='Case Number',
        how='inner'
    )
    merged['source'] = 'BI'
    return merged


def vlookup_ba_with_catalog(ba_df, catalog_df):
    ba_df = ba_df.copy()
    ba_df['individual case basis code number'] = ba_df['individual case basis code number'].astype(str)
    catalog_df = catalog_df.copy()
    catalog_df['Case Number'] = catalog_df['Case Number'].astype(str)
    
    merged = ba_df.merge(
        catalog_df[['Case Number', 'EIS CLIN', 'EIS CLIN Name', 'Verizon Case Description', 'SRE Pricing Element']],
        left_on='individual case basis code number',
        right_on='Case Number',
        how='inner'
    )
    merged['source'] = 'BA'
    return merged


def extract_opdiv_from_cor(cor_email):
    if pd.isna(cor_email) or cor_email == '':
        return ''
    cor_str = str(cor_email).lower()
    if '@' in cor_str:
        domain_part = cor_str.split('@')[1]
        if 'nih.gov' in domain_part:
            return 'NIH'
        elif 'cms.hhs.gov' in domain_part:
            return 'CMS'
        elif 'oig.hhs.gov' in domain_part:
            return 'OIG'
        elif 'fda.hhs.gov' in domain_part:
            return 'FDA'
        elif 'cdc.gov' in domain_part:
            return 'CDC'
        elif 'hhs.gov' in domain_part:
            subdomain = domain_part.split('.')[0]
            return subdomain.upper()
    return ''


def create_active_subscriptions_rows(bi_merged, ba_merged, month_value):
    rows = []
    
    for _, row in bi_merged.iterrows():
        cor_value = row.get('contracting officer representative email address', '')
        opdiv_value = extract_opdiv_from_cor(cor_value)
        new_row = {
            'Verizon Term Start Date': '',
            'Verizon Term End Date': '',
            'Type or Renewal (New Order/Continuous/etc)': '',
            'Customer of Record': '',
            'Qty': row.get('quantity', ''),
            'Month BI/BA': month_value,
            'CLIN': row.get('EIS CLIN', ''),
            'CLIN Name': row.get('EIS CLIN Name', ''),
            'Case Number': row.get('Case Number', ''),
            'Case Description': row.get('Verizon Case Description', ''),
            'PE': row.get('SRE Pricing Element', ''),
            'AHC': row.get('agency hierarchy code', ''),
            'COR': cor_value,
            'OPDIV': opdiv_value
        }
        rows.append(new_row)
    
    for _, row in ba_merged.iterrows():
        new_row = {
            'Verizon Term Start Date': '',
            'Verizon Term End Date': '',
            'Type or Renewal (New Order/Continuous/etc)': '',
            'Customer of Record': '',
            'Qty': row.get('quantity', ''),
            'Month BI/BA': month_value,
            'CLIN': row.get('EIS CLIN', ''),
            'CLIN Name': row.get('EIS CLIN Name', ''),
            'Case Number': row.get('Case Number', ''),
            'Case Description': row.get('Verizon Case Description', ''),
            'PE': row.get('SRE Pricing Element', ''),
            'AHC': row.get('agency hierarchy code', ''),
            'COR': '',
            'OPDIV': ''
        }
        rows.append(new_row)
    
    return pd.DataFrame(rows)


def update_active_subscriptions(existing_active_df, new_rows_df):
    updated_active = pd.concat([existing_active_df, new_rows_df], ignore_index=True)
    return updated_active


def get_month_date_range(month_name, year=2025):
    month_map = {
        'January': (1, 31), 'February': (2, 28), 'March': (3, 31),
        'April': (4, 30), 'May': (5, 31), 'June': (6, 30),
        'July': (7, 31), 'August': (8, 31), 'September': (9, 30),
        'October': (10, 31), 'November': (11, 30), 'December': (12, 31)
    }
    month_num, last_day = month_map.get(month_name, (12, 31))
    start_date = pd.Timestamp(year=year, month=month_num, day=1)
    end_date = pd.Timestamp(year=year, month=month_num, day=last_day)
    return start_date, end_date


def identify_expired_subscriptions(active_df, month_name, year=2025):
    start_date, end_date = get_month_date_range(month_name, year)
    active_df = active_df.copy()
    
    def is_in_date_range(val):
        if pd.isna(val) or val == '' or val == 'Vz Entitled Equipment' or val == 'EXPIRED':
            return False
        try:
            date_val = pd.to_datetime(val, errors='coerce')
            if pd.isna(date_val):
                return False
            return start_date <= date_val <= end_date
        except:
            return False
    
    mask = active_df['Verizon Term End Date'].apply(is_in_date_range)
    expired_rows = active_df[mask].copy()
    expired_row_indices = active_df[mask].index.tolist()
    remaining_active = active_df[~mask].copy()
    
    return expired_rows, remaining_active, expired_row_indices


def format_expired_subscriptions_sheet(ws):
    from openpyxl.styles import Border, Side
    
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True, size=11)
    black_font = Font(color='000000', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    green_headers = ['Verizon Term Start Date', 'Verizon Term End Date', 
                     'Type or Renewal (New Order/Continuous/etc)', 'Customer of Record']
    
    date_columns = []
    for col_idx, cell in enumerate(ws[1], 1):
        header_value = cell.value
        if header_value in ['Verizon Term Start Date', 'Verizon Term End Date']:
            date_columns.append(col_idx)
        if header_value in green_headers:
            cell.fill = green_fill
            cell.font = black_font
        else:
            cell.fill = black_fill
            cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border
            if cell.column in date_columns and cell.value:
                cell.number_format = 'M/D/YYYY'
    
    max_width = 50
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def save_to_excel(catalog_df, active_df, expired_df, original_file, output_path, new_catalog_rows_df, new_active_rows_df, expired_row_indices):
    from openpyxl.styles import Border, Side
    
    wb = load_workbook(original_file)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    ws_catalog = wb['Catalog of Subscription CLINs']
    if len(new_catalog_rows_df) > 0:
        start_row = ws_catalog.max_row + 1
        for r_idx, row in enumerate(new_catalog_rows_df.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_catalog.cell(row=r_idx, column=c_idx, value=value)
                cell.fill = yellow_fill
                cell.border = thin_border
                cell.alignment = left_alignment
    
    ws_active = wb['J.17 Active Subscriptions']
    if len(expired_row_indices) > 0:
        for row_idx in sorted(expired_row_indices, reverse=True):
            ws_active.delete_rows(row_idx + 2)
    
    if len(new_active_rows_df) > 0:
        start_row = ws_active.max_row + 1
        for r_idx, row in enumerate(new_active_rows_df.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_active.cell(row=r_idx, column=c_idx, value=value)
                cell.fill = yellow_fill
                cell.border = thin_border
                cell.alignment = center_alignment
    
    ws_expired = wb['J.17 Expired Subscriptions']
    ws_expired.delete_rows(1, ws_expired.max_row)
    
    headers = list(expired_df.columns)
    ws_expired.append(headers)
    for row in expired_df.itertuples(index=False):
        ws_expired.append(row)
    
    format_expired_subscriptions_sheet(ws_expired)
    
    wb.save(output_path)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Create the updated J.17 workbook from selected J.17, J.1, and billing workbooks."
    )
    parser.add_argument("--j17-file", help="Source J.17 workbook")
    parser.add_argument("--j1-previous-file", help="Baseline J.1 previous workbook")
    parser.add_argument("--j1-current-file", help="Updated/current J.1 workbook")
    parser.add_argument("--billing-file", help="EIS Billing Detail workbook")
    parser.add_argument("--output-file", help="Output updated J.17 workbook path")
    parser.add_argument("--option-period", type=int, default=current_option_period, help="Current option period")
    parser.add_argument("--current-month", default=DEFAULT_CURRENT_MONTH,
                        help="Month used to identify expired subscriptions; defaults to December to preserve original behavior")
    return parser.parse_args(argv)


def configure_paths(args):
    global j17_file, j1_previous_file_input, j1_current_file, eis_billing_file
    global j17_updated_file, current_option_period, current_month
    current_option_period = args.option_period
    j17_file = choose_existing_file(args.j17_file, "Select the source J.17 workbook", EXCEL_FILETYPES)
    j1_previous_file_input = choose_existing_file(args.j1_previous_file, "Select the baseline J.1 previous workbook", EXCEL_FILETYPES)
    j1_current_file = choose_existing_file(args.j1_current_file, "Select the updated/current J.1 workbook", EXCEL_FILETYPES)
    eis_billing_file = choose_existing_file(args.billing_file, "Select the EIS Billing Detail workbook", EXCEL_FILETYPES)
    j17_updated_file = choose_save_file(args.output_file, "Save the updated J.17 workbook as", "j17_updated_file.xlsx", EXCEL_FILETYPES)
    current_month = args.current_month or DEFAULT_CURRENT_MONTH


def main(argv=None):
    args = parse_args(argv)
    configure_paths(args)
    j17_updated_file.parent.mkdir(parents=True, exist_ok=True)

    month_value = extract_month_from_filename(str(eis_billing_file))

    catalog_df, active_df, expired_df = load_existing_j17_data(j17_file)

    sheet_name = get_catalog_sheet_name(current_option_period)
    original_j1_df = read_excel_auto(j1_previous_file_input, sheet_name=sheet_name)
    original_row_count = len(original_j1_df)

    full_j1_df = load_draft_j1_catalog(j1_current_file, current_option_period)
    draft_j1_df = full_j1_df.iloc[original_row_count:].copy()
    print(f"J1 current file: {len(full_j1_df)} total rows, {original_row_count} from prior mods, {len(draft_j1_df)} from current mod")

    new_catalog_rows_df = get_new_catalog_rows(catalog_df, draft_j1_df)

    updated_catalog_df = pd.concat([catalog_df, new_catalog_rows_df], ignore_index=True)

    bi_df, ba_df = load_billing_data(eis_billing_file)

    bi_merged = vlookup_bi_with_catalog(bi_df, updated_catalog_df)
    ba_merged = vlookup_ba_with_catalog(ba_df, updated_catalog_df)

    new_active_rows_df = create_active_subscriptions_rows(bi_merged, ba_merged, month_value)

    expired_rows, remaining_active_df, expired_row_indices = identify_expired_subscriptions(
        active_df, current_month
    )

    final_expired_df = expired_rows.copy()

    save_to_excel(
        updated_catalog_df, remaining_active_df, final_expired_df,
        j17_file, j17_updated_file,
        new_catalog_rows_df, new_active_rows_df, expired_row_indices
    )

    print(f"J.17 updated file created: {j17_updated_file}")


if __name__ == "__main__":
    main()