import pandas as pd
import numpy as np
import os
import re
import argparse
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from file_selection import (
    choose_existing_file, choose_existing_dir, choose_output_dir,
    read_excel_auto, iter_excel_files, EXCEL_FILETYPES,
)


# ─── Runtime-selected inputs ──────────────────────────────────────────────────
# These are assigned in main() from CLI arguments or file/folder picker dialogs.
PR_DIR = None
OUTPUT_COVERSHEETS_DIR = None
j1_previous_file = None
clin_table_file = None


# ─── Configuration ────────────────────────────────────────────────────────────
CURRENT_OP = 5  # Override with --current-op when the period changes.

# ─── Constants ────────────────────────────────────────────────────────────────
COVERSHEET_HEADERS = [
    "OpDiv Approval (Yes/ No)",
    "New CLIN, Rate Reduction, Rate Correction, or Update?",
    "S. No.",
    "CLIN",
    "CLIN Description",
    "Pricing Factor",
    "Originating Address",
    "Terminating Address",
    "SRE Pricing Element",
    "Band Low",
    "Band High",
    "Frequency",
    "QTY",
    "HHS Price per CLIN ($)",
    "TO Period",
    "\u2265 12 Month CLIN",
    "Included in P000# ",
    "MSA Mod Number",
    "MSA Status (Pending, Awarded, or N/A)",
    "F&R Status (Pending, Approved, or N/A)",
]

PRICING_METHOD_COLUMNS = {
    "NONE": ["EIS CLIN", "Band Low", "Band High", "TO Period"],
    "ICB": [
        "EIS CLIN", "Case Number", "SRE Pricing Element",
        "Band Low", "Band High", "TO Period",
    ],
    "ORIG NSC": ["EIS CLIN", "Orig NSC", "Band Low", "Band High", "TO Period"],
    "ORIG NSC-TERM NSC": [
        "EIS CLIN", "Orig NSC", "Term NSC", "Band Low", "Band High", "TO Period",
    ],
    "TERM NSC": ["EIS CLIN", "Term NSC", "Band Low", "Band High", "TO Period"],
    "ORIG JUR": ["EIS CLIN", "Orig CJID", "Band Low", "Band High", "TO Period"],
    "ORIG JUR-TERM JUR": [
        "EIS CLIN", "Orig CJID", "Term CJID", "Band Low", "Band High", "TO Period",
    ],
    "TERM JUR": ["EIS CLIN", "Term CJID", "Band Low", "Band High", "TO Period"],
}


# ─── Utility helpers ──────────────────────────────────────────────────────────

def normalize_value(val):
    """Return a hashable, type-consistent representation of *val*.
    NaN / None  → None
    Float that is a whole number (120033.0) → int
    Strings are stripped.
    """
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    if isinstance(val, float) and val == int(val):
        return int(val)
    if isinstance(val, str):
        return val.strip()
    return val


def make_key(row_dict, cols):
    """Build a hashable tuple from *row_dict* for the given *cols*."""
    return tuple(normalize_value(row_dict.get(c)) for c in cols)


def safe_str(val):
    """Convert *val* to a display string; NaN / None → empty string."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def parse_to_period(val):
    """'OPT PD 4' → 4   (returns None when unparseable)."""
    if pd.isna(val):
        return None
    m = re.search(r"(\d+)", str(val))
    return int(m.group(1)) if m else None


# ─── Data-loading functions ───────────────────────────────────────────────────

def load_j1_data():
    """Load and combine data from the two J1 catalog sheets."""
    sheet_2b = f"2B_Opt Pd {CURRENT_OP} Catalog"
    sheet_2c = f"2C_Opt Pd {CURRENT_OP + 1}-11 Catalog"
    j1_2b = read_excel_auto(j1_previous_file, sheet_name=sheet_2b)
    j1_2c = read_excel_auto(j1_previous_file, sheet_name=sheet_2c)
    print(f"  J1 sheets: '{sheet_2b}' + '{sheet_2c}'")
    return pd.concat([j1_2b, j1_2c], ignore_index=True)


def load_clin_lookup():
    """Return dict  { EIS_CLIN_str : Pricing_Method_str }."""
    clin_df = read_excel_auto(clin_table_file)
    lookup = {}
    for _, row in clin_df.iterrows():
        clin_val = str(row["Clin"]).strip()
        pm_val = str(row["Pricing Method"]).strip()
        lookup[clin_val] = pm_val
    return lookup


def build_j1_lookup(j1_data, clin_lookup):
    """Return dict  { key_tuple : HHS_Price }  for fast existence / price checks."""
    j1_lkp = {}
    for _, row in j1_data.iterrows():
        eis_clin = str(row.get("EIS CLIN", "")).strip()
        pm = clin_lookup.get(eis_clin)
        if pm is None:
            continue
        match_cols = PRICING_METHOD_COLUMNS.get(pm)
        if match_cols is None:
            continue
        key = make_key(row.to_dict(), match_cols)
        if key not in j1_lkp:
            hhs = row.get("HHS Price")
            j1_lkp[key] = float(hhs) if pd.notna(hhs) else 0.0
    return j1_lkp


# ─── PR file helpers ──────────────────────────────────────────────────────────

def get_pr_files():
    """Return sorted PR Excel filenames, excluding generated coversheets/temp files."""
    return [p.name for p in iter_excel_files(Path(PR_DIR), exclude_substrings=("_coversheet", "coversheet"))]


def extract_pr_number(filename):
    """'PR56807.HHS.IHS.J.1.2024.06.03.v02.xlsx' → 'PR56807'."""
    m = re.match(r"(PR\d+)", filename)
    return m.group(1) if m else None


# ─── Row-level logic ──────────────────────────────────────────────────────────

def determine_change_type(j1_lookup, key, pr_hhs_price):
    """Return 'New CLIN' | 'Rate Reduction' | 'Rate Correction' | None."""
    if key not in j1_lookup:
        return "New CLIN"
    j1_price = j1_lookup[key]
    pr_price = float(pr_hhs_price) if pd.notna(pr_hhs_price) else 0.0
    if pr_price < j1_price:
        return "Rate Reduction"
    if pr_price > j1_price:
        return "Rate Correction"
    return None


def determine_pricing_factor(pricing_method, pr_row):
    """Derive the Pricing Factor value from the pricing method and PR row."""
    if pricing_method == "ICB":
        return safe_str(pr_row.get("Case Number"))
    if pricing_method == "ORIG NSC":
        return safe_str(pr_row.get("Orig NSC"))
    if pricing_method == "ORIG JUR":
        return safe_str(pr_row.get("Orig CJID"))
    if pricing_method == "ORIG NSC-TERM NSC":
        orig = safe_str(pr_row.get("Orig NSC"))
        term = safe_str(pr_row.get("Term NSC"))
        return f"{orig}_{term}" if (orig or term) else ""
    if pricing_method == "ORIG JUR-TERM JUR":
        orig = safe_str(pr_row.get("Orig CJID"))
        term = safe_str(pr_row.get("Term CJID"))
        return f"{orig}_{term}" if (orig or term) else ""
    if pricing_method == "TERM NSC":
        return safe_str(pr_row.get("Term NSC"))
    if pricing_method == "TERM JUR":
        return safe_str(pr_row.get("Term CJID"))
    return ""


def determine_fr_status(pricing_method, hhs_price):
    """ICB with positive price → 'Pending'; everything else → 'N/A'."""
    if pricing_method == "ICB" and pd.notna(hhs_price) and hhs_price > 0:
        return "Pending"
    return "N/A"


def build_address(pr_row, prefix):
    """Concatenate the four address columns with ', ' separators."""
    parts = [
        safe_str(pr_row.get(prefix)),
        safe_str(pr_row.get(f"{prefix} - City")),
        safe_str(pr_row.get(f"{prefix} - State")),
        safe_str(pr_row.get(f"{prefix} - Zip Code")),
    ]
    return ", ".join(p for p in parts if p)


def build_coversheet_row(pr_row, change_type, pricing_method, to_period_num):
    """Assemble a single coversheet row dict (QTY and S. No. filled later)."""
    band_low = pr_row.get("Band Low")
    band_high = pr_row.get("Band High")
    hhs_price = pr_row.get("HHS Price")

    return {
        "OpDiv Approval (Yes/ No)": "",
        "New CLIN, Rate Reduction, Rate Correction, or Update?": change_type,
        "S. No.": None,
        "CLIN": safe_str(pr_row.get("EIS CLIN")),
        "CLIN Description": safe_str(pr_row.get("EIS CLIN Name")),
        "Pricing Factor": determine_pricing_factor(pricing_method, pr_row),
        "Originating Address": build_address(pr_row, "Originating Address"),
        "Terminating Address": build_address(pr_row, "Terminating Address"),
        "SRE Pricing Element": safe_str(pr_row.get("SRE Pricing Element")),
        "Band Low": float(band_low) if pd.notna(band_low) else "",
        "Band High": float(band_high) if pd.notna(band_high) else "",
        "Frequency": safe_str(pr_row.get("EIS Frequency")),
        "QTY": 1,
        "HHS Price per CLIN ($)": float(hhs_price) if pd.notna(hhs_price) else "",
        "TO Period": to_period_num,
        "\u2265 12 Month CLIN": "",
        "Included in P000# ": "",
        "MSA Mod Number": "",
        "MSA Status (Pending, Awarded, or N/A)": "",
        "F&R Status (Pending, Approved, or N/A)": determine_fr_status(
            pricing_method, hhs_price
        ),
    }


# ─── QTY computation ─────────────────────────────────────────────────────────

def compute_qty_and_deduplicate(rows):
    """Group identical rows → set QTY = group size, assign sequential S. No."""
    if not rows:
        return []

    group_cols = [
        h for h in COVERSHEET_HEADERS if h not in ("S. No.", "QTY")
    ]

    def _row_key(row):
        return tuple(str(row.get(c, "")) for c in group_cols)

    counts = Counter()
    first_seen = {}
    for row in rows:
        k = _row_key(row)
        counts[k] += 1
        if k not in first_seen:
            first_seen[k] = row

    result = []
    for serial, (k, row) in enumerate(first_seen.items(), start=1):
        row["S. No."] = serial
        row["QTY"] = counts[k]
        result.append(row)
    return result


# ─── Processing one PR file ──────────────────────────────────────────────────

def process_pr_file(pr_filepath, j1_lookup, clin_lookup):
    """Return (current_op_rows, oy_rows) for one PR file."""
    pr_data = read_excel_auto(pr_filepath)

    current_op_rows = []
    oy_rows = []

    for _, pr_row in pr_data.iterrows():
        eis_clin = pr_row.get("EIS CLIN")
        if pd.isna(eis_clin):
            continue
        eis_clin_str = str(eis_clin).strip()

        pricing_method = clin_lookup.get(eis_clin_str)
        if pricing_method is None:
            print(
                f"    WARNING: EIS CLIN '{eis_clin_str}' not found "
                "in CLIN table - skipping row"
            )
            continue

        match_cols = PRICING_METHOD_COLUMNS.get(pricing_method)
        if match_cols is None:
            print(
                f"    WARNING: Unrecognized pricing method "
                f"'{pricing_method}' for CLIN '{eis_clin_str}' - skipping row"
            )
            continue

        to_period_num = parse_to_period(pr_row.get("TO Period"))
        if to_period_num is None:
            continue

        pr_dict = pr_row.to_dict()
        pr_dict["TO Period"] = to_period_num

        key = make_key(pr_dict, match_cols)

        change_type = determine_change_type(
            j1_lookup, key, pr_row.get("HHS Price")
        )
        if change_type is None:
            continue

        cs_row = build_coversheet_row(
            pr_row, change_type, pricing_method, to_period_num
        )

        if to_period_num == CURRENT_OP:
            current_op_rows.append(cs_row)
        elif CURRENT_OP < to_period_num <= 11:
            oy_rows.append(cs_row)

    current_op_rows = compute_qty_and_deduplicate(current_op_rows)
    oy_rows = compute_qty_and_deduplicate(oy_rows)

    return current_op_rows, oy_rows


# ─── Excel output ─────────────────────────────────────────────────────────────

def write_sheet(ws, rows, styles):
    """Write headers + data rows into *ws* with formatting."""
    header_font, cell_font, header_align, cell_align, border, yellow_fill = styles

    for col_idx, header in enumerate(COVERSHEET_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        if header == "OpDiv Approval (Yes/ No)":
            cell.fill = yellow_fill

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(COVERSHEET_HEADERS, 1):
            val = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = border

    for col_idx in range(1, len(COVERSHEET_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, len(rows) + 2):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)


def create_coversheet(output_path, current_op_rows, oy_rows):
    """Create a formatted two-sheet coversheet workbook."""
    wb = Workbook()

    ws_current = wb.active
    ws_current.title = "CLIN Table (Current OP)"
    ws_oy = wb.create_sheet("CLIN Table (OY)")

    styles = (
        Font(name="Calibri", size=11, bold=True),
        Font(name="Calibri", size=11),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
        Alignment(horizontal="left", vertical="center"),
        Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    )

    write_sheet(ws_current, current_op_rows, styles)
    write_sheet(ws_oy, oy_rows, styles)

    wb.save(output_path)
    print(f"  Saved -> {os.path.basename(output_path)}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Generate coversheets from selected J.1, CLIN lookup, and PR files."
    )
    parser.add_argument("--j1-previous-file", help="Path to the J.1 previous workbook")
    parser.add_argument("--clin-table-file", help="Path to the CLIN lookup workbook")
    parser.add_argument("--pr-dir", help="Folder containing PR workbooks")
    parser.add_argument("--output-dir", help="Folder where generated coversheets should be saved")
    parser.add_argument("--current-op", type=int, default=CURRENT_OP, help="Current option period")
    return parser.parse_args(argv)


def configure_paths(args):
    global PR_DIR, OUTPUT_COVERSHEETS_DIR, j1_previous_file, clin_table_file, CURRENT_OP
    CURRENT_OP = args.current_op
    j1_previous_file = choose_existing_file(args.j1_previous_file, "Select the J.1 previous workbook", EXCEL_FILETYPES)
    clin_table_file = choose_existing_file(args.clin_table_file, "Select the CLIN lookup workbook", EXCEL_FILETYPES)
    PR_DIR = choose_existing_dir(args.pr_dir, "Select the folder containing PR workbooks")
    OUTPUT_COVERSHEETS_DIR = choose_output_dir(args.output_dir, "Select/create the folder for generated coversheets")


def main(argv=None):
    args = parse_args(argv)
    configure_paths(args)

    print("Loading J1 data ...")
    j1_data = load_j1_data()
    print(f"  {len(j1_data)} rows from 2B + 2C")

    print("Loading CLIN table ...")
    clin_lookup = load_clin_lookup()
    print(f"  {len(clin_lookup)} CLINs")

    print("Building J1 lookup index ...")
    j1_lookup = build_j1_lookup(j1_data, clin_lookup)
    print(f"  {len(j1_lookup)} indexed entries")

    pr_files = get_pr_files()
    print(f"\nFound {len(pr_files)} PR file(s) in {PR_DIR}\n")

    for pr_file in pr_files:
        print(f"Processing: {pr_file}")
        pr_path = Path(PR_DIR) / pr_file
        pr_number = extract_pr_number(pr_file)

        if not pr_number:
            print("  Could not extract PR number - skipping")
            continue

        current_op_rows, oy_rows = process_pr_file(
            pr_path, j1_lookup, clin_lookup
        )

        if not current_op_rows and not oy_rows:
            print("  Both sheets empty -- no coversheet needed\n")
            continue

        coversheet_path = Path(OUTPUT_COVERSHEETS_DIR) / f"{pr_number}_coversheet.xlsx"
        create_coversheet(coversheet_path, current_op_rows, oy_rows)
        print(
            f"  Current OP: {len(current_op_rows)} row(s) | "
            f"OY: {len(oy_rows)} row(s)\n"
        )

    print("All coversheets generated successfully!")


if __name__ == "__main__":
    main()
