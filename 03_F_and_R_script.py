import pandas as pd
import os
import argparse
import re
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
from collections import defaultdict
import math
from file_selection import (
    choose_existing_file, choose_existing_dir, choose_save_file,
    iter_excel_files, read_excel_auto, excel_file, EXCEL_FILETYPES,
)


# ─── Configuration ────────────────────────────────────────────────────────────
CURRENT_OPTION_PERIOD = 5

# ─── Runtime-selected inputs ──────────────────────────────────────────────────
# Assigned in main() from CLI arguments or file/folder picker dialogs.
PR_DIR = None
overview_file = None
f_r_output_file = None


def _clean(val):
    """Normalize value to a clean string; blanks / NaN / 'nan' / 'none' become ''."""
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "<na>"):
        return ""
    return s


_HEADER_PUNCT_RE = re.compile(r"[^a-z0-9]+")
_NUMERIC_TEXT_RE = re.compile(r"^\d+(?:\.0+)?$")


def _normalise_header_name(value):
    """Normalize a worksheet header for tolerant comparisons while reading Row 1 only."""
    s = _clean(value).lower()
    s = s.replace("’", "'").replace("`", "'")
    # Common misspelling found in some COMP templates.
    s = s.replace("reponse", "response")
    s = s.replace("&", "and")
    return _HEADER_PUNCT_RE.sub("", s)


def _is_comment_header(header):
    """Return True for accepted Verizon/HHS comment column-name variants."""
    h = _normalise_header_name(header)
    if not h:
        return False
    return (
        ("verizon" in h and "response" in h and "comment" in h)
        or ("verizon" in h and "response" in h)
        or ("hhs" in h and "comment" in h)
    )


def _comment_columns(df):
    """Find all COMP-sheet columns that should feed the F&R comment output field."""
    return [col for col in df.columns if _is_comment_header(col)]


def _is_source_or_networx_header(header):
    """Return True for accepted Source/Networx/Network information headers.

    This intentionally matches header labels, not cell values. It accepts the
    canonical project wording plus common variants such as "Source Information",
    "Networx Information", "Network Information", and shortened "Info" forms.
    """
    h = _normalise_header_name(header)
    if not h:
        return False

    direct_aliases = {
        _normalise_header_name(alias)
        for alias in (
            "Source or Networx Information",
            "Source or Network Information",
            "Source Information",
            "Networx Information",
            "Network Information",
            "Source or Networx Info",
            "Source or Network Info",
            "Source Info",
            "Networx Info",
            "Network Info",
        )
    }
    if h in direct_aliases:
        return True

    has_info_word = "information" in h or h.endswith("info") or "info" in h
    has_source_word = "source" in h
    has_network_word = "networx" in h or "network" in h

    return has_info_word and (has_source_word or has_network_word)


def _source_or_networx_columns(df):
    """Find COMP-sheet columns that should feed Source or Networx Information."""
    return [col for col in df.columns if _is_source_or_networx_header(col)]


def _find_column_by_alias(df, aliases):
    """Return the first DataFrame column matching any normalized alias."""
    normalized_to_column = {_normalise_header_name(col): col for col in df.columns}
    for alias in aliases:
        col = normalized_to_column.get(_normalise_header_name(alias))
        if col is not None:
            return col
    return None


def _get_row_value_by_alias(df, row_index, aliases):
    """Get a row value from the first matching header alias, or blank if none exists."""
    col = _find_column_by_alias(df, aliases)
    if col is None:
        return ""
    try:
        return df[col].iloc[row_index]
    except Exception:
        return ""


def _combined_row_text(df, row_index, columns):
    """Combine unique nonblank values from one or more source columns in a row."""
    values = []
    for col in columns:
        try:
            value = _clean(df[col].iloc[row_index])
        except Exception:
            value = ""
        if value and value not in values:
            values.append(value)
    return " / ".join(values)


def _normalise_case_number(value):
    """Normalize case numbers for matching without requiring identical Excel cell types."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            if float(value).is_integer():
                return str(int(value))
        except (ValueError, TypeError):
            pass
    s = _clean(value)
    if not s:
        return ""
    s = s.replace("\u00a0", " ").strip()
    if _NUMERIC_TEXT_RE.fullmatch(s):
        # Convert values like 123.0 to 123 while preserving true text values like 00123.
        if "." in s:
            return s.split(".", 1)[0]
        return s
    return re.sub(r"\s+", "", s).upper()


def _case_numbers_match(left, right):
    left_norm = _normalise_case_number(left)
    right_norm = _normalise_case_number(right)
    if left_norm == right_norm:
        return True
    if left_norm.isdigit() and right_norm.isdigit():
        try:
            return int(left_norm) == int(right_norm)
        except ValueError:
            return False
    return False


def _normalise_pricing_element(value):
    """Normalize SRE Pricing Element values for matching/output."""
    s = _clean(value)
    if not s:
        return ""
    s = s.replace("\u00a0", " ").strip()
    try:
        numeric_value = float(s)
        if numeric_value.is_integer():
            return str(int(numeric_value)).zfill(2)
    except (ValueError, TypeError):
        pass
    return s.zfill(2)


def _pricing_elements_match(left, right):
    left_norm = _normalise_pricing_element(left)
    right_norm = _normalise_pricing_element(right)
    if left_norm == right_norm:
        return True
    try:
        return int(float(left_norm)) == int(float(right_norm))
    except (ValueError, TypeError):
        return False


def _normalise_description(value):
    """Normalize case-description text for J.1 tie-break matching."""
    s = _clean(value).upper().replace("\u00a0", " ")
    return re.sub(r"[^A-Z0-9]+", "", s)


def _descriptions_match(left, right):
    left_norm = _normalise_description(left)
    right_norm = _normalise_description(right)
    return bool(left_norm and right_norm and left_norm == right_norm)


def _unique_price_or_none(candidates):
    """Return a candidate price only if all nonblank candidate prices agree."""
    prices = []
    for candidate in candidates:
        price = candidate.get("price")
        if price is not None:
            try:
                rounded = round(float(price), 9)
            except (ValueError, TypeError):
                continue
            if rounded not in prices:
                prices.append(rounded)
    if len(prices) == 1:
        return candidates[0].get("price")
    return None


def _option_period_number(value):
    """Extract the option-period number from common TO Period text/value formats."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            if float(value).is_integer():
                return int(value)
        except (ValueError, TypeError):
            return None
    s = _clean(value).upper()
    if not s:
        return None
    numbers = re.findall(r"\d+", s)
    if not numbers:
        return None
    try:
        return int(numbers[-1])
    except ValueError:
        return None


def _option_period_matches(value, current_option_period):
    return _option_period_number(value) == int(current_option_period)


def _parse_currency_value(value):
    """Parse numeric/currency values from J.1 or COMP cells."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = _clean(value)
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = text.replace("$", "").replace(",", "").strip()
    try:
        parsed = float(text)
        return -parsed if negative else parsed
    except (ValueError, TypeError):
        return None


def get_fr_pr_numbers(overview_path: Path):
    """
    Reads the overview file and returns a list of tuples for rows where 'F&R Needed' == 'Yes'.
    Each tuple contains:
        (PR#, Version, OpDiv, SF30 Description, 12M+ CLIN)
    """
    # Read Excel and strip column names
    df = read_excel_auto(overview_path, dtype=str)  # read all as string to preserve formatting
    df.columns = df.columns.str.strip()

    # Filter rows where F&R Needed == 'Yes'
    fr_needed = df[df["F&R Needed"].astype(str).str.strip().str.lower() == "yes"]

    # Strip whitespace and get required columns
    pr_version_list = list(
        zip(
            fr_needed["PR#"].map(_clean),
            fr_needed["Version"].map(_clean),
            fr_needed["OpDiv"].map(_clean) if "OpDiv" in fr_needed.columns else [""] * len(fr_needed),
            fr_needed["SF30 Description"].map(_clean) if "SF30 Description" in fr_needed.columns else [""] * len(fr_needed),
            fr_needed["12M+ CLIN"].map(_clean) if "12M+ CLIN" in fr_needed.columns else [""] * len(fr_needed),
        )
    )

    print(f"Found {len(pr_version_list)} PRs needing F&R.")
    return pr_version_list


def extract_comps_data(comps_df):
    """
    Extract all relevant F&R fields from the Comps worksheet in order.

    The output workbook keeps the canonical column name
    "Verizon's Response/HHS Comment", but the source COMP sheet may use
    variants such as "Verizon Response", "Verizon's Reponse",
    "Verizon's Response", or "HHS Comment". When separate Verizon and HHS
    comment columns exist, their nonblank values are combined.
    """
    records = []
    num_rows = len(comps_df)
    comment_cols = _comment_columns(comps_df)
    if not comment_cols:
        print("  No Verizon/HHS response/comment column variant found on COMP sheet")

    source_cols = _source_or_networx_columns(comps_df)
    if not source_cols:
        print("  No Source/Networx/Network information column variant found on COMP sheet")

    for i in range(num_rows):
        verizon_response = _combined_row_text(comps_df, i, comment_cols)
        source_info = _combined_row_text(comps_df, i, source_cols)
        case_number = _get_row_value_by_alias(comps_df, i, ["Case Number", "Case #", "Case No"])
        verizon_case_desc = _get_row_value_by_alias(
            comps_df, i,
            ["Verizon Case Description", "Case Description", "ICB Case Description", "Description"],
        )
        pricing_element = _get_row_value_by_alias(
            comps_df, i,
            [
                "SRE Pricing Element",
                "Pricing Element",
                "Element Number",
                "Element #",
                "Element No",
                "Element",
            ],
        )
        comp_rate_raw = _get_row_value_by_alias(comps_df, i, ["Comp Rate", "Comparison Rate", "Comparable Rate"])
        comp_rate = _parse_currency_value(comp_rate_raw)

        record = {
            "Verizon's Response/HHS Comment": _clean(verizon_response),
            "Source or Networx Information": _clean(source_info),
            "Case Number": _clean(case_number),
            "Verizon Case Description": _clean(verizon_case_desc),
            "Pricing Element": _clean(pricing_element),
            "Comp Rate": comp_rate,
        }

        records.append(record)

    return records


def format_currency(value):
    """Format a numeric value as currency with 6 decimal places."""
    if value is None:
        return ""
    try:
        return f"${float(value):,.6f}"
    except (ValueError, TypeError):
        return ""


def get_comps_sheet(pr_file: Path):
    """Load the 'comps' worksheet (case-insensitive) and return as DataFrame."""
    try:
        xls = excel_file(pr_file)
        comps_sheet = next((s for s in xls.sheet_names if "comp" in s.lower()), None)
        if comps_sheet:
            df = pd.read_excel(xls, sheet_name=comps_sheet)
            df.columns = df.columns.str.strip()
            return df
        else:
            print(f"  No 'comps' sheet found in {pr_file.name}")
            return pd.DataFrame()
    except Exception as e:
        print(f"  Error reading comps sheet from {pr_file.name}: {e}")
        return pd.DataFrame()



def determine_12m_clin(verizon_case_desc):
    """
    Determine 12M+ CLIN based on Verizon Case Description.
    Returns 'Yes' if '.ANN.' is found in the description, otherwise 'No'.
    """
    if verizon_case_desc is None:
        return "No"
    
    desc_str = str(verizon_case_desc).strip()
    return "Yes" if ".ANN." in desc_str.upper() else "No"



def _select_j1_candidate(candidates, target_case, target_pricing, target_desc, pr_file_name):
    """Select the safest J.1 match from candidates already filtered by case/period."""
    if not candidates:
        print(
            f"  No J.1 price match in {pr_file_name} "
            f"for Case Number={target_case}, SRE Pricing Element={target_pricing or '[blank]'}, "
            f"TO Period=OPT PD {CURRENT_OPTION_PERIOD}"
        )
        return None

    # If COMP supplied a pricing element, those rows are already filtered by element.
    # Return the first matching price to preserve the original behavior.
    if target_pricing:
        return candidates[0].get("price")

    # If COMP did not supply a pricing element, use the case description as the
    # tie-breaker. This handles PR rows with the same Case Number and blank SRE
    # Pricing Element but separate NRC/MRC J.1 lines.
    if target_desc:
        desc_matches = [c for c in candidates if _descriptions_match(c.get("description"), target_desc)]
        if desc_matches:
            return desc_matches[0].get("price")

    if len(candidates) == 1:
        return candidates[0].get("price")

    same_price = _unique_price_or_none(candidates)
    if same_price is not None:
        return same_price

    print(
        f"  Ambiguous J.1 price match in {pr_file_name} "
        f"for Case Number={target_case}, blank SRE Pricing Element, "
        f"TO Period=OPT PD {CURRENT_OPTION_PERIOD}; leaving J.1 Rate blank"
    )
    return None


def get_j1_rate(pr_file: Path, case_number: str, pricing_element: str, case_description: str = ""):
    """
    Extract J.1 rate from the J.1 worksheet for a given case number and pricing element.

    J.1 headers are expected on Row 1 of the PR file. Matching is tolerant to
    Excel cell-type differences. If the COMP row has a blank pricing element and
    multiple J.1 rows share the same case/period, the COMP case description is
    used as a tie-breaker against the J.1 ICB Case Description.
    """
    target_case = _normalise_case_number(case_number)
    target_pricing = _normalise_pricing_element(pricing_element)
    target_desc = _clean(case_description)

    try:
        if pr_file.suffix.lower() == ".xlsb":
            xls = excel_file(pr_file)
            j1_sheet = next((s for s in xls.sheet_names if "j1" in s.lower() or "j.1" in s.lower()), None)
            if not j1_sheet:
                print(f"  No 'J1' sheet found in {pr_file.name}")
                return None
            df = pd.read_excel(xls, sheet_name=j1_sheet)
            df.columns = [str(c).strip() for c in df.columns]
            required = {"Case Number", "SRE Pricing Element", "TO Period", "HHS Price"}
            if not required.issubset(set(df.columns)):
                print(f"  Missing required Row 1 columns in J1 sheet of {pr_file.name}")
                return None
            candidates = []
            for _, row in df.iterrows():
                if not _case_numbers_match(row.get("Case Number"), target_case):
                    continue
                if not _option_period_matches(row.get("TO Period"), CURRENT_OPTION_PERIOD):
                    continue
                if target_pricing and not _pricing_elements_match(row.get("SRE Pricing Element"), target_pricing):
                    continue
                candidates.append({
                    "price": _parse_currency_value(row.get("HHS Price")),
                    "description": row.get("ICB Case Description", ""),
                })
            return _select_j1_candidate(candidates, target_case, target_pricing, target_desc, pr_file.name)

        wb = load_workbook(pr_file, data_only=True)
        j1_sheet = next((s for s in wb.sheetnames if "j1" in s.lower() or "j.1" in s.lower()), None)
        if not j1_sheet:
            print(f"  No 'J1' sheet found in {pr_file.name}")
            wb.close()
            return None

        ws = wb[j1_sheet]

        # Headers are intentionally read from Row 1 only.
        headers = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = idx

        required = {"Case Number", "SRE Pricing Element", "TO Period", "HHS Price"}
        if not required.issubset(set(headers)):
            print(f"  Missing required Row 1 columns in J1 sheet of {pr_file.name}")
            wb.close()
            return None

        case_col = headers["Case Number"]
        pricing_col = headers["SRE Pricing Element"]
        period_col = headers["TO Period"]
        price_col = headers["HHS Price"]
        desc_col = headers.get("ICB Case Description")

        candidates = []
        for row in ws.iter_rows(min_row=2):
            row_case = row[case_col - 1].value
            row_pricing = row[pricing_col - 1].value
            row_period = row[period_col - 1].value
            if not _case_numbers_match(row_case, target_case):
                continue
            if not _option_period_matches(row_period, CURRENT_OPTION_PERIOD):
                continue
            if target_pricing and not _pricing_elements_match(row_pricing, target_pricing):
                continue
            row_description = row[desc_col - 1].value if desc_col else ""
            candidates.append({
                "price": _parse_currency_value(row[price_col - 1].value),
                "description": row_description,
            })

        result = _select_j1_candidate(candidates, target_case, target_pricing, target_desc, pr_file.name)
        wb.close()
        return result

    except Exception as e:
        print(f"  Error reading J1 sheet from {pr_file.name}: {e}")
        return None


def build_FR():
    """
    Builds the F&R Overview Excel file replicating the official example formatting,
    converts Pricing Element to numeric when possible, and merges duplicate Case Numbers.
    Also creates separate PR tabs before merging.
    """
    f_r_output_file.parent.mkdir(parents=True, exist_ok=True)

    # Step 1: Get PRs needing F&R
    fr_pr_list = get_fr_pr_numbers(overview_file)
    fr_overview_records = []
    pr_to_file = {}

    # Step 2: Gather data
    for pr, version, opdiv, sf30_desc, cl12m in fr_pr_list:
        pr_identifier = pr if "PR" in pr.upper() else f"PR{pr}"
        print(f"Processing {pr_identifier}...")

        pr_files = [
            f for f in iter_excel_files(PR_DIR)
            if f.name.upper().startswith(pr_identifier.upper())
        ]
        if not pr_files:
            print(f"  No PR files found for {pr_identifier} in {PR_DIR}")
            continue

        for pr_file in pr_files:
            comps_df = get_comps_sheet(pr_file)
            if comps_df.empty:
                continue

            if pr not in pr_to_file:
                pr_to_file[pr] = pr_file

            comps_records = extract_comps_data(comps_df)

            for rec in comps_records:
                case_num = rec.get("Case Number", "")
                case_num_str = _normalise_case_number(case_num)

                price_elem_raw = rec.get("Pricing Element", "")
                price_elem_str = _normalise_pricing_element(price_elem_raw)

                j1_rate_float = get_j1_rate(pr_file, case_num_str, price_elem_str, rec.get("Verizon Case Description", ""))
                comp_rate_float = rec.get("Comp Rate")

                delta_float = None
                if j1_rate_float is not None and comp_rate_float is not None:
                    delta_float = j1_rate_float - comp_rate_float

                j1_rate = format_currency(j1_rate_float)
                comp_rate = format_currency(comp_rate_float)
                delta = format_currency(delta_float)

                fr_overview_records.append({
                    "Type of F&R": "",
                    "Verizon's Response/HHS Comment": rec.get("Verizon's Response/HHS Comment", ""),
                    "J.1 Rate": j1_rate,
                    "Comp Rate": comp_rate,
                    "Delta": delta,
                    "Source or Networx Information": rec.get("Source or Networx Information", ""),
                    "Case Number": rec.get("Case Number", ""),
                    "Verizon Case Description": rec.get("Verizon Case Description", ""),
                    "Pricing Element": price_elem_str,
                    "PR#": pr,
                    "Version": str(version).zfill(2),
                    "OpDiv": opdiv,
                    "SF30 Description": sf30_desc,
                    "12M+ CLIN": determine_12m_clin(rec.get("Verizon Case Description", ""))
                })

    if not fr_overview_records:
        print("No F&R records found.")
        return None

    # Create workbook early (before merging) to build PR tabs
    wb = Workbook()

    # --- Shared Styles ---
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    black_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # --- Define columns ---
    columns = [
        "Type of F&R",
        "Verizon's Response/HHS Comment",
        "J.1 Rate",
        "Comp Rate",
        "Delta",
        "Source or Networx Information",
        "Case Number",
        "Verizon Case Description",
        "Pricing Element",
        "PR#",
        "Version",
        "OpDiv",
        "SF30 Description",
        "12M+ CLIN"
    ]

    # Group F&R records by PR for later use in xlwings PR tab creation
    pr_groups = defaultdict(list)
    for rec in fr_overview_records:
        pr_groups[rec["PR#"]].append(rec)

    # Remove default "Sheet"
    default_ws = wb.active
    wb.remove(default_ws)

    # Merge duplicates and create Overview tab
    merged_records = []
    merge_key_columns = ("Case Number", "Pricing Element")
    concat_columns = {
        "PR#", "Version", "OpDiv", "SF30 Description",
        "12M+ CLIN", "Verizon's Response/HHS Comment", "Source or Networx Information"
    }

    temp_dict = defaultdict(list)
    for record in fr_overview_records:
        key = (record["Case Number"], record["Pricing Element"])
        temp_dict[key].append(record)

    for (case_number, pricing_element), records in temp_dict.items():
        merged_record = {}
        for col in records[0].keys():
            if col in merge_key_columns:
                merged_record[col] = records[0][col]
            elif col in concat_columns:
                values = [str(r[col]).strip() for r in records if str(r[col]).strip()]
                unique_values = list(dict.fromkeys(values))
                merged_record[col] = "/".join(unique_values) if unique_values else ""
            else:
                values = [str(r[col]).strip() for r in records if str(r[col]).strip()]
                merged_record[col] = max(values, key=len) if values else ""
        merged_records.append(merged_record)

    fr_overview_records = merged_records

    # Create Overview tab
    ws = wb.create_sheet(title="Overview")
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if col_idx <= 6:
            cell.fill = green_fill
            cell.font = black_font
        elif 7 <= col_idx <= 9:
            cell.fill = black_fill
            cell.font = white_font
        elif 10 <= col_idx <= 11:
            cell.fill = gray_fill
            cell.font = black_font
        elif 12 <= col_idx <= 13:
            cell.fill = yellow_fill
            cell.font = black_font
        elif col_idx == 14:
            cell.fill = light_green_fill
            cell.font = black_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, record in enumerate(fr_overview_records, start=2):
        for col_idx, header in enumerate(columns, start=1):
            value = record.get(header, "")
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            if 10 <= col_idx <= 11:
                c.fill = gray_fill
            elif 12 <= col_idx <= 13:
                c.fill = yellow_fill
            elif col_idx == 14:
                c.fill = light_green_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, min(max_len + 3, 45))
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(columns)).coordinate}"
    ws.freeze_panes = "A2"

    wb.save(f_r_output_file)
    print(f"\nOverview tab created. Now adding PR tabs with images via xlwings...")

    # Step 3: Use xlwings to create PR tabs by copying COMP sheets from PR files
    _create_pr_tabs_xlwings(pr_groups, pr_to_file, columns[:9], f_r_output_file)

    print(f"\nF&R Overview document created with PR tabs and merged duplicates: {f_r_output_file}")
    return f_r_output_file


def _create_pr_tabs_xlwings(pr_groups, pr_to_file, fr_headers, output_file):
    """
    For each PR needing F&R:
    1. Copy the COMP sheet from the PR file into f_r_output as a new tab
    2. Delete the original Comps header + data rows
    3. Insert F&R header and data rows at the top
    Images in the COMP sheet are preserved automatically.
    """
    import xlwings as xw

    currency_cols = {"J.1 Rate", "Comp Rate", "Delta"}
    currency_fmt = '$#,##0.000000'

    def _parse_currency(val):
        """Strip '$' and commas to recover the raw float."""
        if not val or not isinstance(val, str):
            return val
        stripped = val.replace("$", "").replace(",", "").strip()
        if not stripped:
            return ""
        try:
            return round(float(stripped), 6)
        except (ValueError, TypeError):
            return val

    app = xw.App(visible=False, add_book=False)
    try:
        wb_output = app.books.open(str(output_file))

        for pr, records in pr_groups.items():
            pr_file = pr_to_file.get(pr)
            if not pr_file:
                print(f"  No PR file found for {pr}, skipping PR tab")
                continue

            pr_version = records[0]["Version"]
            tab_name = f"{pr} v{pr_version}"
            print(f"  Creating tab: {tab_name}")

            # Open PR file and find Comps sheet
            wb_pr = app.books.open(str(pr_file), read_only=True)
            comps_sheet = None
            for sht in wb_pr.sheets:
                if "comp" in sht.name.lower():
                    comps_sheet = sht
                    break

            if comps_sheet is None:
                print(f"    No COMP sheet found in {pr_file.name}")
                wb_pr.close()
                continue

            # Copy the COMP sheet to the output workbook
            comps_sheet.copy(after=wb_output.sheets[-1])
            wb_pr.close()

            # Rename the copied sheet
            new_sheet = wb_output.sheets[-1]
            new_sheet.name = tab_name

            # Find the end of the original Comps data (first fully empty row)
            used = new_sheet.used_range
            comps_data_end = 0
            for row_num in range(1, used.rows.count + 1):
                row_vals = new_sheet.range((row_num, 1), (row_num, used.columns.count)).value
                if row_vals is None:
                    break
                if isinstance(row_vals, list):
                    if all(v is None for v in row_vals):
                        break
                comps_data_end = row_num

            # Delete the original Comps data rows (header + data)
            if comps_data_end > 0:
                new_sheet.range(f"1:{comps_data_end}").delete()

            # Insert blank rows at the top for F&R header + data
            num_fr_rows = 1 + len(records)  # 1 header + N data rows
            new_sheet.range(f"1:{num_fr_rows}").insert(shift='down')

            # Write F&R header row
            for col_idx, header in enumerate(fr_headers):
                cell = new_sheet.range((1, col_idx + 1))
                cell.value = header
                cell.font.bold = True
                if col_idx < 6:
                    cell.color = (146, 208, 80)  # green
                else:
                    cell.color = (0, 0, 0)  # black
                    cell.font.color = (255, 255, 255)  # white text

            # Write F&R data rows
            for row_idx, record in enumerate(records, start=2):
                for col_idx, header in enumerate(fr_headers):
                    value = record.get(header, "")
                    cell = new_sheet.range((row_idx, col_idx + 1))
                    if header in currency_cols:
                        cell.value = _parse_currency(value)
                        if cell.value != "" and cell.value is not None:
                            cell.number_format = currency_fmt
                    else:
                        cell.value = value

            # Apply borders to F&R header + data range
            last_row = 1 + len(records)
            last_col = len(fr_headers)
            fr_range = new_sheet.range((1, 1), (last_row, last_col))
            for border_id in range(7, 13):
                fr_range.api.Borders(border_id).LineStyle = 1  # xlContinuous
                fr_range.api.Borders(border_id).Weight = 2     # xlThin

            # Auto-fit columns
            new_sheet.autofit('c')

        wb_output.save()
        wb_output.close()
    finally:
        app.quit()


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Build the F&R workbook from a selected overview file and PR folder."
    )
    parser.add_argument("--overview-file", help="Overview workbook path")
    parser.add_argument("--pr-dir", help="Folder containing PR workbooks")
    parser.add_argument("--output-file", help="Output F&R workbook path")
    parser.add_argument("--current-option-period", type=int, default=CURRENT_OPTION_PERIOD,
                        help="Current option period used for J.1 rate matching")
    return parser.parse_args(argv)


def configure_paths(args):
    global overview_file, PR_DIR, f_r_output_file, CURRENT_OPTION_PERIOD
    CURRENT_OPTION_PERIOD = args.current_option_period
    overview_file = choose_existing_file(args.overview_file, "Select the overview workbook", EXCEL_FILETYPES)
    PR_DIR = choose_existing_dir(args.pr_dir, "Select the folder containing PR workbooks")
    f_r_output_file = choose_save_file(args.output_file, "Save the F&R workbook as", "f_r_output.xlsx", EXCEL_FILETYPES)


def main(argv=None):
    args = parse_args(argv)
    configure_paths(args)
    return build_FR()


if __name__ == "__main__":
    main()
