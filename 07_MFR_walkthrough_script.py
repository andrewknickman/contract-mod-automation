"""
07_MFR_walkthrough_script.py

Reads the J1 current file (produced by 05_J1_script.py) and generates the MFR
Walkthrough Output Word document by counting rows in three catalog sheets based
on yellow highlighting and column criteria.

Dependencies: openpyxl, python-docx
    pip install openpyxl python-docx
"""

import os
import argparse
from pathlib import Path

from openpyxl import load_workbook
from docx import Document
from docx.shared import Pt, Inches
from file_selection import (
    choose_existing_file, choose_save_file, EXCEL_FILETYPES, WORD_FILETYPES,
)


# ─── Configuration ────────────────────────────────────────────────────────────
option_period = 5
mod_number = None


# ─── Runtime-selected inputs ──────────────────────────────────────────────────
# Assigned in main() from CLI arguments or file picker dialogs.
j1_current_file = None
mfr_walkthrough_output = None


# ─── Constants ────────────────────────────────────────────────────────────────
YELLOW_RGB = "FFFF00"


# ============================================================
# SHEET / HIGHLIGHT HELPERS
# ============================================================

def get_sheet_names(opt_pd):
    """Return the three relevant sheet names for a given option period."""
    sheet_2b = f"2B_Opt Pd {opt_pd} Catalog"
    sheet_2c = f"2C_Opt Pd {opt_pd + 1}-11 Catalog"
    sheet_2d = "2D_Deleted Items"
    return sheet_2b, sheet_2c, sheet_2d


def is_yellow_fill(cell):
    """True iff `cell` has a solid fill whose RGB ends with FFFF00."""
    fill = cell.fill
    if fill is None or fill.patternType != 'solid':
        return False
    fg = fill.fgColor
    if fg is None or fg.type != 'rgb' or not fg.rgb:
        return False
    # openpyxl stores fgColor.rgb as ARGB ('FFFFFF00') or RGB ('FFFF00')
    return str(fg.rgb).upper().endswith(YELLOW_RGB)


def find_header_columns(ws):
    """Return list of stripped header strings for the contiguous header block in row 1."""
    headers = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        headers.append(str(v).strip() if v is not None else "")
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def find_column_index(headers, target_name):
    """1-based index of a header (case-insensitive, stripped match), or None."""
    target = target_name.strip().lower()
    for idx, h in enumerate(headers, 1):
        if h.strip().lower() == target:
            return idx
    return None


def analyze_sheet(ws):
    """
    Walk every non-blank data row of `ws` and return two parallel lists:
      rows_data        — list of dicts mapping header -> cell value
      highlight_flags  — list of dicts with booleans 'all_yellow' and 'hhs_yellow'

    Trailing blank rows (where every used column is empty) are skipped.
    """
    headers = find_header_columns(ws)
    n_cols = len(headers)
    hhs_col_idx_1 = find_column_index(headers, "HHS Price")
    hhs_col_idx_0 = hhs_col_idx_1 - 1 if hhs_col_idx_1 else None

    rows_data = []
    highlight_flags = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=n_cols):
        values = [cell.value for cell in row]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
            continue

        all_yellow = all(is_yellow_fill(cell) for cell in row)
        hhs_yellow = (
            is_yellow_fill(row[hhs_col_idx_0])
            if hhs_col_idx_0 is not None
            else False
        )

        rows_data.append(dict(zip(headers, values)))
        highlight_flags.append({'all_yellow': all_yellow, 'hhs_yellow': hhs_yellow})

    return rows_data, highlight_flags


# ============================================================
# VALUE COERCION
# ============================================================

def to_float(v, default=None):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return default


def to_int(v, default=None):
    f = to_float(v, default=None)
    if f is None:
        return default
    return int(f)


def get_field(row, name):
    val = row.get(name, "")
    return "" if val is None else val


# ============================================================
# COUNT COMPUTATION
# ============================================================

def compute_counts_2b(rows, flags):
    """Return all counts derived from the 2B sheet: b, C1, C2, d, e, f, g, h, i, j..n."""
    c1 = c2 = 0
    c2_rows = []
    for row, hl in zip(rows, flags):
        if hl['all_yellow']:
            c2 += 1
            c2_rows.append(row)
        elif hl['hhs_yellow']:
            c1 += 1
    b = c1 + c2

    d = 0
    e_rows = []
    for r in c2_rows:
        pm = str(get_field(r, 'YY_Price_Method')).strip().upper()
        if pm == 'ICB':
            e_rows.append(r)
        else:
            d += 1
    e = len(e_rows)

    f = 0
    g_rows = []
    for r in e_rows:
        price = to_float(get_field(r, 'HHS Price'), default=0.0)
        if price == 0:
            f += 1
        else:
            g_rows.append(r)
    g = len(g_rows)

    h = 0
    i_rows = []
    for r in g_rows:
        clin = str(get_field(r, 'EIS CLIN')).strip().upper()
        if clin.startswith('EQ'):
            i_rows.append(r)
        else:
            h += 1
    i_count = len(i_rows)

    pe_counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for r in i_rows:
        pe = to_int(get_field(r, 'SRE Pricing Element'))
        if pe in pe_counts:
            pe_counts[pe] += 1

    return {
        'b': b, 'C1': c1, 'C2': c2,
        'd': d, 'e': e, 'f': f, 'g': g,
        'h': h, 'i': i_count,
        'j': pe_counts[1], 'k': pe_counts[2], 'l': pe_counts[3],
        'm': pe_counts[4], 'n': pe_counts[5],
    }


def compute_counts_2c(rows, flags):
    """Return all counts derived from the 2C sheet: o, P1, P2, q, r, s, t, u, v."""
    p1 = p2 = 0
    p2_rows = []
    for row, hl in zip(rows, flags):
        if hl['all_yellow']:
            p2 += 1
            p2_rows.append(row)
        elif hl['hhs_yellow']:
            p1 += 1
    o = p1 + p2

    q = 0
    r_rows = []
    for r in p2_rows:
        pm = str(get_field(r, 'YY_Price_Method')).strip().upper()
        if pm == 'ICB':
            r_rows.append(r)
        else:
            q += 1
    r_count = len(r_rows)

    s = 0
    t_rows = []
    for r in r_rows:
        price = to_float(get_field(r, 'HHS Price'), default=0.0)
        if price == 0:
            s += 1
        else:
            t_rows.append(r)
    t = len(t_rows)

    u = v_count = 0
    for r in t_rows:
        clin = str(get_field(r, 'EIS CLIN')).strip().upper()
        if clin.startswith('EQ'):
            v_count += 1
        else:
            u += 1

    return {
        'o': o, 'P1': p1, 'P2': p2,
        'q': q, 'r': r_count, 's': s, 't': t,
        'u': u, 'v': v_count,
    }


def compute_counts_2d(rows, flags):
    """w = rows where HHS Price-only is yellow + rows where entire row is yellow."""
    w = sum(1 for hl in flags if hl['all_yellow'] or hl['hhs_yellow'])
    return {'w': w}


# ============================================================
# DOCX GENERATION
# ============================================================

def add_para(doc, text, indent_level=0):
    p = doc.add_paragraph(text)
    p.paragraph_format.left_indent = Inches(0.5 * indent_level)
    for run in p.runs:
        run.font.name = "Calibri"
        run.font.size = Pt(11)
    return p


def generate_walkthrough_doc(values, mod_number, option_period, output_path):
    """Create the MFR Walkthrough Output Word document at `output_path`."""
    sheet_2b, sheet_2c, sheet_2d = get_sheet_names(option_period)
    doc = Document()

    add_para(doc, f"{mod_number} touches {values['a']} items.", 0)
    add_para(doc, f"{values['b']} items were added or updated on Tab {sheet_2b} (Column W on {mod_number}).", 1)
    add_para(doc, f"{values['C1']} new line items were updated on Tab {sheet_2b}.", 2)
    add_para(doc, f"{values['C2']} new line items were added on Tab {sheet_2b}.", 2)
    add_para(doc, f"{values['d']} were standard CLINs (Column A Unselect ICB).", 3)
    add_para(doc, f"{values['e']} were ICBs (Column A on ICB only).", 3)
    add_para(doc, f"{values['f']} were zero-dollar ICBs and did not require F&R.", 4)
    add_para(doc, f"{values['g']} ICBs required F&R.", 4)
    add_para(doc, f"{values['h']} Non-SRE (Filter B on all but EQ)", 5)
    add_para(doc, f"{values['i']} SRE (Filter B on EQ)", 5)
    add_para(doc, f"{values['j']} PE 01 \u2013 Published List Price", 6)
    add_para(doc, f"{values['k']} PE 02 \u2013 Parametric", 6)
    add_para(doc, f"{values['l']} PE 03 \u2013 Published List Price", 6)
    add_para(doc, f"{values['m']} PE 04 \u2013 Parametric", 6)
    add_para(doc, f"{values['n']} PE 05 \u2013 Published List Price", 6)
    add_para(doc, f"{values['o']} items were added or updated on Tab {sheet_2c} (Column W on {mod_number}).", 1)
    add_para(doc, f"{values['P1']} new line items were updated on Tab {sheet_2c}.", 2)
    add_para(doc, f"{values['P2']} new line items were added on Tab {sheet_2c}.", 2)
    add_para(doc, f"{values['q']} were standard CLINs (Column A Unselect ICB).", 3)
    add_para(doc, f"{values['r']} were ICBs (Column A on ICB only).", 3)
    add_para(doc, f"{values['s']} were zero dollar ICBs and did not require F&R.", 4)
    add_para(doc, f"{values['t']} ICBs required F&R.", 4)
    add_para(doc, f"{values['u']} Non-SRE (Filter B on all but EQ)", 5)
    add_para(doc, f"{values['v']} SRE (Filter B on EQ)", 5)
    add_para(doc, f"{values['w']} items were added or updated on Tab {sheet_2d} (Column W on {mod_number}).", 1)

    doc.save(output_path)


# ============================================================
# MAIN
# ============================================================

def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Generate the MFR Walkthrough Word document from a selected J.1 current workbook."
    )
    parser.add_argument("--j1-current-file", help="Updated/current J.1 workbook")
    parser.add_argument("--output-file", help="Output Word document path")
    parser.add_argument("--option-period", type=int, default=option_period, help="Option period to analyze")
    parser.add_argument("--mod-number", required=False, help="Mod number for the document heading/output default")
    return parser.parse_args(argv)


def configure_paths(args):
    global j1_current_file, mfr_walkthrough_output, option_period, mod_number
    option_period = args.option_period
    mod_number = args.mod_number or input("Mod number for MFR Walkthrough document, for example P00078: ").strip()
    default_name = f"{mod_number} MFR Walkthrough Output.docx" if mod_number else "MFR Walkthrough Output.docx"
    j1_current_file = choose_existing_file(args.j1_current_file, "Select the updated/current J.1 workbook", EXCEL_FILETYPES)
    mfr_walkthrough_output = choose_save_file(args.output_file, "Save the MFR Walkthrough document as", default_name, WORD_FILETYPES)


def main(argv=None):
    args = parse_args(argv)
    configure_paths(args)
    mfr_walkthrough_output.parent.mkdir(parents=True, exist_ok=True)

    sheet_2b_name, sheet_2c_name, sheet_2d_name = get_sheet_names(option_period)

    print("=" * 60)
    print("MFR WALKTHROUGH GENERATION")
    print("=" * 60)
    print(f"Mod Number     : {mod_number}")
    print(f"Option Period  : {option_period}")
    print(f"Input file     : {j1_current_file}")
    print(f"Output file    : {mfr_walkthrough_output}")
    print(f"Sheets         : {sheet_2b_name} | {sheet_2c_name} | {sheet_2d_name}")
    print("=" * 60)

    if not j1_current_file.exists():
        raise FileNotFoundError(f"J1 current file not found: {j1_current_file}")

    wb = load_workbook(j1_current_file, data_only=True)
    try:
        ws_2b = wb[sheet_2b_name]
        ws_2c = wb[sheet_2c_name]
        ws_2d = wb[sheet_2d_name]

        print(f"\nReading {sheet_2b_name}...")
        rows_2b, flags_2b = analyze_sheet(ws_2b)
        print(f"  {len(rows_2b)} data rows")

        print(f"Reading {sheet_2c_name}...")
        rows_2c, flags_2c = analyze_sheet(ws_2c)
        print(f"  {len(rows_2c)} data rows")

        print(f"Reading {sheet_2d_name}...")
        rows_2d, flags_2d = analyze_sheet(ws_2d)
        print(f"  {len(rows_2d)} data rows")
    finally:
        wb.close()

    print("\nComputing counts...")
    counts_2b = compute_counts_2b(rows_2b, flags_2b)
    counts_2c = compute_counts_2c(rows_2c, flags_2c)
    counts_2d = compute_counts_2d(rows_2d, flags_2d)

    a = counts_2b['b'] + counts_2c['o'] + counts_2d['w']
    values = {'a': a, **counts_2b, **counts_2c, **counts_2d}

    print("\nComputed values:")
    for key in ['a', 'b', 'C1', 'C2', 'd', 'e', 'f', 'g', 'h', 'i',
                'j', 'k', 'l', 'm', 'n', 'o', 'P1', 'P2', 'q', 'r',
                's', 't', 'u', 'v', 'w']:
        print(f"  {key:>3} = {values[key]}")

    print(f"\nGenerating Word doc: {mfr_walkthrough_output}")
    generate_walkthrough_doc(values, mod_number, option_period, mfr_walkthrough_output)

    print("\n" + "=" * 60)
    print(f"MFR Walkthrough output created: {mfr_walkthrough_output}")
    print("=" * 60)


if __name__ == "__main__":
    main()
