#!/usr/bin/env python3
"""
Combined automation script for Build File creation.
This script:
1. Copies the build file from input/ to output/ (preserving the original)
2. Processes PR and Coversheet files to create J.1 Automated sheet
3. Creates Catalog sheet from J.1 data
"""

import pandas as pd
import re
import shutil
import argparse
from pathlib import Path
import warnings
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from file_selection import (
    choose_existing_file, choose_existing_dir, choose_save_file,
    iter_excel_files, find_excel_by_prefix, read_excel_auto, excel_file,
    EXCEL_FILETYPES,
)

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore')


# ─── Runtime-selected inputs ──────────────────────────────────────────────────
# Assigned in main() from CLI arguments or file/folder picker dialogs.
PR_DIR = None
COVERSHEET_FILES_DIR = None
build_file_input = None
build_file_output = None


# ============================================================
# SECTION 1: J.1 SHEET AUTOMATION FUNCTIONS
# ============================================================

def extract_pr_number_from_coversheet(filename: str) -> Optional[str]:
    """
    Extract PR number from coversheet filename.
    
    Args:
        filename: Coversheet filename (e.g., "TO Mod Coversheet MOD-0004567-NIH_PR60203.xlsx")
        
    Returns:
        PR number or None if not found
    """
    # Look for pattern: PR followed by digits (after any prefix like NIH_, IHS_, etc.)
    match = re.search(r'PR(\d+)', filename, re.IGNORECASE)
    if match:
        return f"PR{match.group(1)}"
    return None


def find_matching_pr_file(pr_number: str, pr_path: Path) -> Optional[Path]:
    """
    Find the corresponding PR file for a given PR number.
    
    Args:
        pr_number: The PR number to search for
        pr_path: Path to the PR files directory
        
    Returns:
        Path to the matching PR file or None if not found
    """
    if not pr_path.exists():
        print(f"Error: PR directory does not exist: {pr_path}")
        return None
        
    # Search for an Excel PR file that starts with the PR number.
    # Supports .xlsx, .xlsm, .xlsb, and .xls instead of hard-coding .xlsx.
    return find_excel_by_prefix(pr_path, pr_number)


def convert_to_period(to_period_value) -> Optional[str]:
    """
    Convert TO Period from coversheet format to PR format.
    
    Args:
        to_period_value: TO Period value from coversheet (e.g., '5' or 5)
        
    Returns:
        Formatted TO Period for PR matching (e.g., 'OPT PD 5')
    """
    if pd.isna(to_period_value):
        return None
        
    # Convert to string and strip whitespace
    to_period_str = str(to_period_value).strip()
    
    # If it's a number, format it as 'OPT PD X'
    if to_period_str.isdigit():
        return f"OPT PD {to_period_str}"
    
    # If it already contains 'OPT PD', return as is
    if 'OPT PD' in to_period_str.upper():
        return to_period_str
        
    return to_period_str


def load_pricing_method_mapping(build_file_path: Path) -> Dict[str, str]:
    """
    Load the CLIN to Pricing Method mapping from the build file.
    
    Args:
        build_file_path: Path to the P00070 Build.xlsx file
        
    Returns:
        Dictionary mapping EIS CLIN to Pricing Method
    """
    pricing_map = {}
    
    try:
        # Read the first CLIN Table-like sheet instead of a date-stamped sheet name.
        xls = excel_file(build_file_path)
        clin_sheet = next((sheet for sheet in xls.sheet_names if 'clin table' in sheet.lower()), None)
        if not clin_sheet:
            raise ValueError("No sheet containing 'CLIN Table' was found in the build workbook")
        clin_table = pd.read_excel(xls, sheet_name=clin_sheet)
        
        # Create mapping from Clin to Pricing Method
        for _, row in clin_table.iterrows():
            if pd.notna(row.get('Clin')) and pd.notna(row.get('Pricing Method')):
                pricing_map[str(row['Clin']).strip()] = str(row['Pricing Method']).strip()
        
        print(f"Loaded {len(pricing_map)} CLIN to Pricing Method mappings")
        
    except Exception as e:
        print(f"Error loading pricing method mapping: {str(e)}")
    
    return pricing_map


def format_date_column(date_value):
    """
    Convert date from '2026-08-31 0:00:00' format to 'DD-MM-YYYY' format.
    
    Args:
        date_value: Date value in various formats
        
    Returns:
        Formatted date string in DD-MM-YYYY format or original value if not a date
    """
    if pd.isna(date_value):
        return ""
    
    try:
        if isinstance(date_value, str):
            date_obj = pd.to_datetime(date_value)
        elif isinstance(date_value, pd.Timestamp):
            date_obj = date_value
        else:
            return date_value
        
        return date_obj.strftime('%d-%m-%Y')
        
    except:
        return date_value


def format_dataframe_dates(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format date columns in the dataframe before writing to Excel.
    
    Args:
        df: DataFrame with date columns
        
    Returns:
        DataFrame with formatted date columns
    """
    df_copy = df.copy()
    
    # List of date columns to format
    date_columns = ['Price Start Date', 'Price End Date']
    
    for col in date_columns:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(format_date_column)
    
    return df_copy


# Coversheet schema helpers. The Build step is often run against a shared folder
# that contains both TO Mod Coversheets and PR/J.1 workbooks. These helpers keep
# PR workbooks from being misread as coversheets and prevent exact header-spacing
# differences from crashing the build process.
COVERSHEET_CURRENT_OP_SHEET_ALIASES = (
    "CLIN Table (Current OP)",
    "CLIN Table Current OP",
    "CLIN Table - Current OP",
)
COVERSHEET_OY_SHEET_ALIASES = (
    "CLIN Table (OY)",
    "CLIN Table OY",
    "CLIN Table - OY",
)
OPDIV_APPROVAL_COLUMN_ALIASES = (
    "OpDiv Approval (Yes/ No)",
    "OpDiv Approval (Yes/No)",
    "OpDiv Approval (Yes / No)",
    "OpDiv Approval Yes No",
    "OpDiv Approval (Y/N)",
    "OpDiv Approval Y/N",
    "OpDiv Approval",
    "OpDiv Approved",
    "OPDIV Approval",
)
TO_PERIOD_COLUMN_ALIASES = (
    "TO Period",
    "TO Periods",
    "Period",
)


def normalise_schema_label(value) -> str:
    """Normalize workbook headers/sheet names for tolerant schema matching."""
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def find_column_by_alias(df: pd.DataFrame, aliases: Tuple[str, ...]) -> Optional[str]:
    """Return the actual DataFrame column matching one of the accepted aliases."""
    wanted = {normalise_schema_label(alias) for alias in aliases}
    for column in df.columns:
        if normalise_schema_label(column) in wanted:
            return column
    return None


def find_sheet_by_alias(xls, aliases: Tuple[str, ...]) -> Optional[str]:
    """Return the actual workbook sheet name matching one of the accepted aliases."""
    wanted = {normalise_schema_label(alias) for alias in aliases}
    for sheet_name in xls.sheet_names:
        if normalise_schema_label(sheet_name) in wanted:
            return sheet_name
    return None


def workbook_has_coversheet_clin_table(path: Path) -> bool:
    """True when the workbook appears to be a TO Mod Coversheet workbook."""
    try:
        xls = excel_file(path)
        return bool(
            find_sheet_by_alias(xls, COVERSHEET_CURRENT_OP_SHEET_ALIASES)
            or find_sheet_by_alias(xls, COVERSHEET_OY_SHEET_ALIASES)
        )
    except Exception as exc:
        print(f"  Skipping unreadable Excel file {path.name}: {exc}")
        return False


def normalize_pricing_method(value) -> str:
    """Normalize a CLIN Table pricing method label for pricing-factor matching."""
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip().upper()
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"\s*[-]+\s*", "-", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_pricing_factor_value(value) -> str:
    """Normalize pricing-factor values so 120036, 120036.0, and '120036' compare alike."""
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "null"}:
        return ""
    text = text.replace(",", "")
    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except (ValueError, TypeError):
        pass
    return text


def combine_pricing_factor_parts(*values) -> Optional[str]:
    """Combine two pricing-factor values with a hyphen when both are available."""
    normalized = [normalize_pricing_factor_value(value) for value in values]
    if all(normalized):
        return "-".join(normalized)
    return None


def get_pricing_factor_value(pr_row: pd.Series, pricing_method: str) -> Optional[str]:
    """
    Get the appropriate PR/J.1 pricing-factor value based on the CLIN Table
    pricing method. The pricing methods are business labels; the PR/J.1 source
    columns for JUR methods are still named Orig CJID / Term CJID.
    
    Supported pricing methods:
    - ICB
    - ORIG NSC
    - TERM NSC
    - ORIG JUR
    - TERM JUR
    - ORIG JUR-TERM JUR
    - ORIG NSC-TERM NSC
    """
    pricing_method = normalize_pricing_method(pricing_method)

    if pricing_method == 'ICB':
        return pr_row.get('Case Number')
    if pricing_method == 'ORIG NSC':
        return pr_row.get('Orig NSC')
    if pricing_method == 'TERM NSC':
        return pr_row.get('Term NSC')
    if pricing_method == 'ORIG JUR':
        return pr_row.get('Orig CJID')
    if pricing_method == 'TERM JUR':
        return pr_row.get('Term CJID')
    if pricing_method == 'ORIG NSC-TERM NSC':
        return combine_pricing_factor_parts(pr_row.get('Orig NSC'), pr_row.get('Term NSC'))
    if pricing_method == 'ORIG JUR-TERM JUR':
        return combine_pricing_factor_parts(pr_row.get('Orig CJID'), pr_row.get('Term CJID'))

    return None


def read_coversheet_data(coversheet_file: Path) -> pd.DataFrame:
    """
    Read and combine coversheet CLIN data from available coversheet sheets.
    
    Args:
        coversheet_file: Path to the coversheet file
        
    Returns:
        Combined DataFrame with Current OP and/or OY rows
    """
    try:
        xls = excel_file(coversheet_file)
        sheet_names = []
        for aliases in (COVERSHEET_CURRENT_OP_SHEET_ALIASES, COVERSHEET_OY_SHEET_ALIASES):
            sheet_name = find_sheet_by_alias(xls, aliases)
            if sheet_name and sheet_name not in sheet_names:
                sheet_names.append(sheet_name)

        if not sheet_names:
            print(
                f"Error reading coversheet {coversheet_file.name}: no CLIN Table Current OP/OY sheet found"
            )
            return pd.DataFrame()

        frames = []
        for sheet_name in sheet_names:
            frames.append(read_excel_auto(coversheet_file, sheet_name=sheet_name))

        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    except Exception as e:
        print(f"Error reading coversheet {coversheet_file.name}: {str(e)}")
        return pd.DataFrame()


def _normalise_approval_value(value) -> str:
    """Normalize an OpDiv approval cell for inclusion decisions."""
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip().lower()
    if text in {"", "nan", "none", "null"}:
        return ""
    return re.sub(r"[^a-z0-9]+", "", text)


def _is_blank_like(value) -> bool:
    """True when a coversheet cell should be treated as blank for row screening."""
    if value is None or pd.isna(value):
        return True
    return str(value).strip().lower() in {"", "nan", "none", "null"}


def _filter_valid_clin_rows(coversheet_data: pd.DataFrame, to_period_col: str) -> pd.DataFrame:
    """Return rows that have enough coversheet data to attempt a PR/J.1 match."""
    required_cols = []
    if "CLIN" in coversheet_data.columns:
        required_cols.append("CLIN")
    else:
        print("  Warning: CLIN column not found in coversheet data.")
        return pd.DataFrame()

    if to_period_col in coversheet_data.columns:
        required_cols.append(to_period_col)

    valid_mask = pd.Series(True, index=coversheet_data.index)
    for column in required_cols:
        valid_mask &= ~coversheet_data[column].apply(_is_blank_like)

    valid_rows = coversheet_data[valid_mask].copy()
    skipped = len(coversheet_data) - len(valid_rows)
    if skipped:
        print(f"  Skipped {skipped} row(s) without a usable CLIN/TO Period")

    return valid_rows


def filter_approved_items(coversheet_data: pd.DataFrame) -> pd.DataFrame:
    """
    Select coversheet rows for Build processing.

    Business rule:
    - If the coversheet contains explicit OpDiv decisions, include only rows marked Yes.
    - If the coversheet has no explicit Yes/No decisions, include all valid CLIN rows
      as pending-review build candidates. Blank/unknown approval status is not treated
      as rejected unless the coversheet also contains explicit Yes/No decisions.
    
    Args:
        coversheet_data: Combined coversheet DataFrame
        
    Returns:
        DataFrame containing rows eligible for Build matching
    """
    if coversheet_data.empty:
        return pd.DataFrame()

    to_period_col = find_column_by_alias(coversheet_data, TO_PERIOD_COLUMN_ALIASES)
    if not to_period_col:
        print("  Warning: TO Period column not found in coversheet data.")
        return pd.DataFrame()

    valid_rows = _filter_valid_clin_rows(coversheet_data, to_period_col)
    if valid_rows.empty:
        return pd.DataFrame()

    approval_col = find_column_by_alias(valid_rows, OPDIV_APPROVAL_COLUMN_ALIASES)
    if not approval_col:
        print(
            "  Warning: OpDiv approval column not found. "
            "Including all valid CLIN rows as pending-review build candidates."
        )
        included_items = valid_rows.copy()
    else:
        normalized_approvals = valid_rows[approval_col].apply(_normalise_approval_value)
        explicit_decision_mask = normalized_approvals.isin({"yes", "y", "no", "n"})

        if explicit_decision_mask.any():
            yes_mask = normalized_approvals.isin({"yes", "y"})
            included_items = valid_rows[yes_mask].copy()
            print(
                f"  OpDiv approval decisions found: including {len(included_items)} Yes row(s) "
                f"and excluding {len(valid_rows) - len(included_items)} non-Yes row(s)."
            )
        else:
            unknown_count = sum(1 for value in normalized_approvals if value)
            if unknown_count:
                print(
                    f"  Warning: OpDiv approval column has {unknown_count} nonblank value(s) "
                    "but no explicit Yes/No decisions. Including all valid CLIN rows as pending-review build candidates."
                )
            else:
                print(
                    "  No explicit OpDiv Yes/No decisions found. "
                    "Including all valid CLIN rows as pending-review build candidates."
                )
            included_items = valid_rows.copy()
    
    # Add converted TO Period for matching. Keep the canonical helper column name
    # so downstream matching logic remains unchanged.
    included_items['TO_Period_Match'] = included_items[to_period_col].apply(convert_to_period)
    
    return included_items


def read_pr_data(pr_file: Path) -> pd.DataFrame:
    """
    Read PR data from the J1 sheet.
    
    Args:
        pr_file: Path to the PR file
        
    Returns:
        DataFrame with PR J1 data
    """
    try:
        return read_excel_auto(pr_file, sheet_name='J1')
    except Exception as e:
        print(f"Error reading PR file {pr_file.name}: {str(e)}")
        return pd.DataFrame()


def match_coversheet_to_pr_with_pricing(approved_items: pd.DataFrame, 
                                       pr_data: pd.DataFrame,
                                       pricing_map: Dict[str, str]) -> pd.DataFrame:
    """
    Match approved coversheet items to PR data considering pricing methods.
    
    Args:
        approved_items: DataFrame with approved coversheet items
        pr_data: DataFrame with PR J1 data
        pricing_map: Dictionary mapping CLIN to Pricing Method
        
    Returns:
        DataFrame with matched PR rows
    """
    if approved_items.empty or pr_data.empty:
        return pd.DataFrame()
    
    matched_rows = []
    
    for _, coversheet_row in approved_items.iterrows():
        # Get basic matching conditions
        to_period_match = coversheet_row['TO_Period_Match']
        clin_match = coversheet_row['CLIN']
        description_match = coversheet_row['CLIN Description']
        price_match_raw = coversheet_row['HHS Price per CLIN ($)']
        try:
            price_match = round(float(price_match_raw), 2)
        except (ValueError, TypeError):
            price_match = 0.0
        
        # Get pricing method for this CLIN
        pricing_method = pricing_map.get(str(clin_match).strip())
        
        # Base matching conditions
        base_conditions = (
            (pr_data['TO Period'] == to_period_match) &
            (pr_data['EIS CLIN'] == clin_match) &
            (pr_data['EIS CLIN Name'] == description_match) &
            (pr_data['HHS Price'].round(2) == price_match)
        )
        
        # If there's a pricing method that requires additional matching
        if pricing_method and 'Pricing Factor' in coversheet_row.index:
            pricing_factor_coversheet = coversheet_row.get('Pricing Factor')
            
            if pd.notna(pricing_factor_coversheet):
                # Create a mask for pricing factor matching
                pr_data_copy = pr_data.copy()
                pr_data_copy['Pricing_Factor_Match'] = pr_data_copy.apply(
                    lambda row: normalize_pricing_factor_value(get_pricing_factor_value(row, pricing_method)), axis=1
                )
                
                # Add pricing factor condition
                pricing_conditions = (
                    base_conditions & 
                    (pr_data_copy['Pricing_Factor_Match'] == normalize_pricing_factor_value(pricing_factor_coversheet))
                )
                
                pr_matches = pr_data[pricing_conditions]
            else:
                # No pricing factor in coversheet, use base conditions
                pr_matches = pr_data[base_conditions]
        else:
            # No pricing method or no Pricing Factor column, use base conditions
            pr_matches = pr_data[base_conditions]
        
        if not pr_matches.empty:
            matched_rows.append(pr_matches)
        else:
            # Try with less strict matching (without price)
            base_conditions_no_price = (
                (pr_data['TO Period'] == to_period_match) &
                (pr_data['EIS CLIN'] == clin_match) &
                (pr_data['EIS CLIN Name'] == description_match)
            )
            
            # Apply same pricing factor logic without price
            if pricing_method and 'Pricing Factor' in coversheet_row.index:
                pricing_factor_coversheet = coversheet_row.get('Pricing Factor')
                
                if pd.notna(pricing_factor_coversheet):
                    pr_data_copy = pr_data.copy()
                    pr_data_copy['Pricing_Factor_Match'] = pr_data_copy.apply(
                        lambda row: normalize_pricing_factor_value(get_pricing_factor_value(row, pricing_method)), axis=1
                    )
                    
                    pricing_conditions = (
                        base_conditions_no_price & 
                        (pr_data_copy['Pricing_Factor_Match'] == normalize_pricing_factor_value(pricing_factor_coversheet))
                    )
                    
                    pr_matches = pr_data[pricing_conditions]
                else:
                    pr_matches = pr_data[base_conditions_no_price]
            else:
                pr_matches = pr_data[base_conditions_no_price]
            
            if not pr_matches.empty:
                matched_rows.append(pr_matches)
    
    if matched_rows:
        result_df = pd.concat(matched_rows, ignore_index=True)
        # Remove duplicates if any
        result_df = result_df.drop_duplicates()
        return result_df
    else:
        return pd.DataFrame()


def process_single_coversheet(coversheet_file: Path, pr_path: Path, pricing_map: Dict[str, str]) -> pd.DataFrame:
    """
    Process a single coversheet file and match it with corresponding PR data.
    
    Args:
        coversheet_file: Path to the coversheet file
        pr_path: Path to the PR files directory
        pricing_map: Dictionary mapping CLIN to Pricing Method
        
    Returns:
        DataFrame with matched PR data
    """
    print(f"Processing coversheet: {coversheet_file.name}")
    
    # Extract PR number from coversheet filename
    pr_number = extract_pr_number_from_coversheet(coversheet_file.name)
    if not pr_number:
        print(f"Warning: Could not extract PR number from: {coversheet_file.name}")
        return pd.DataFrame()
    
    print(f"  Extracted PR number: {pr_number}")
    
    # Find matching PR file
    pr_file = find_matching_pr_file(pr_number, pr_path)
    if not pr_file:
        print(f"  Warning: Could not find PR file for PR number: {pr_number}")
        return pd.DataFrame()
    
    print(f"  Found matching PR file: {pr_file.name}")
    
    # Read coversheet data
    coversheet_data = read_coversheet_data(coversheet_file)
    
    # Select rows eligible for build matching. If no explicit OpDiv decisions exist,
    # blank approval status is treated as pending review and all valid CLIN rows are included.
    approved_items = filter_approved_items(coversheet_data)
    
    if approved_items.empty:
        print(f"  No build-eligible items found in {coversheet_file.name}")
        return pd.DataFrame()
    
    # Read PR data
    pr_data = read_pr_data(pr_file)
    
    # Match coversheet to PR data with pricing method consideration
    matched_data = match_coversheet_to_pr_with_pricing(approved_items, pr_data, pricing_map)
    
    return matched_data


def get_coversheet_files(coversheet_path: Path) -> List[Path]:
    """
    Get list of all Excel files in the coversheet directory.
    
    Args:
        coversheet_path: Path to the coversheet directory
        
    Returns:
        List of Path objects for Excel files
    """
    if not coversheet_path.exists():
        print(f"Error: Coversheet directory does not exist: {coversheet_path}")
        return []
    
    excel_files = list(iter_excel_files(coversheet_path))
    coversheet_files = [path for path in excel_files if workbook_has_coversheet_clin_table(path)]
    skipped = len(excel_files) - len(coversheet_files)
    print(f"\nFound {len(coversheet_files)} coversheet files to process")
    if skipped:
        print(f"Skipped {skipped} non-coversheet Excel files in the selected coversheet folder")
    return coversheet_files


def process_all_coversheets(coversheet_path: Path, pr_path: Path, build_file_path: Path) -> pd.DataFrame:
    """
    Process all coversheet files and compile matched data.
    
    Args:
        coversheet_path: Path to the coversheet directory
        pr_path: Path to the PR files directory
        build_file_path: Path to the P00070 Build.xlsx file
        
    Returns:
        Combined DataFrame with all matched data
    """
    # Load pricing method mapping
    pricing_map = load_pricing_method_mapping(build_file_path)
    
    # Get all coversheet files
    coversheet_files = get_coversheet_files(coversheet_path)
    
    if not coversheet_files:
        print(f"No Excel files found in {coversheet_path}")
        return pd.DataFrame()
    
    # Process each coversheet and collect results
    all_matched_data = []
    
    for coversheet_file in coversheet_files:
        matched_data = process_single_coversheet(coversheet_file, pr_path, pricing_map)
        if not matched_data.empty:
            all_matched_data.append(matched_data)
    
    # Combine all matched data
    if all_matched_data:
        combined_data = pd.concat(all_matched_data, ignore_index=True)
        # Remove any duplicates across all processed files
        combined_data = combined_data.drop_duplicates()
        print(f"\nTotal rows collected: {len(combined_data)}")
        return combined_data
    else:
        print("No data was matched from any coversheet files")
        return pd.DataFrame()


def format_excel_sheet(worksheet, num_rows, num_cols, dataframe):
    """
    Apply formatting to the Excel worksheet.
    
    Args:
        worksheet: openpyxl worksheet object
        num_rows: Number of data rows (excluding header)
        num_cols: Number of columns
        dataframe: The DataFrame to get column names
    """
    # Define styles
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    sla_header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    red_font = Font(color="FF0000", size=11)
    
    # Get column indices
    column_names = dataframe.columns.tolist()
    sla_col_index = None
    hhs_price_col_index = None
    
    # Find the column indices
    for idx, col_name in enumerate(column_names, 1):
        if "Standard EIS CLIN SLA" in col_name and "Y/N" in col_name:
            sla_col_index = idx
        if col_name == "HHS Price":
            hhs_price_col_index = idx
    
    # Format headers (first row)
    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        
        # Apply special fill for SLA column, normal fill for others
        if col == sla_col_index:
            cell.fill = sla_header_fill
        else:
            cell.fill = header_fill
            
        cell.alignment = center_alignment
    
    # Format data rows
    for row in range(2, num_rows + 2):  # +2 because row 1 is header, and we need to include all data rows
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply red font to HHS Price column values
            if col == hhs_price_col_index:
                cell.font = red_font
    
    # Auto-fit columns with max width
    for col in range(1, num_cols + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        # Check all cells in the column
        for row in range(1, num_rows + 2):
            cell = worksheet.cell(row=row, column=col)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Adjust width with a max limit
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        worksheet.column_dimensions[column_letter].width = adjusted_width


def write_to_build_file(processed_data: pd.DataFrame, build_file_path: Path) -> bool:
    """
    Write the processed data to the J.1 Automated sheet in the build file with formatting.
    
    Args:
        processed_data: DataFrame with all matched data
        build_file_path: Path to the build file
        
    Returns:
        True if successful, False otherwise
    """
    if processed_data.empty:
        print("No data to write to build file")
        return False
    
    try:
        processed_data = format_dataframe_dates(processed_data)
        # Check if build file exists
        if build_file_path.exists():
            # Load existing workbook
            workbook = load_workbook(build_file_path, keep_links=True, data_only=False)
            
            # Remove 'J.1 Automated' sheet if it exists
            if 'J.1 Automated' in workbook.sheetnames:
                del workbook['J.1 Automated']
            
            # Create new sheet
            worksheet = workbook.create_sheet('J.1 Automated')
        else:
            # Create new workbook
            workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # Create new sheet
            worksheet = workbook.create_sheet('J.1 Automated')
        
        # Write headers
        for col_idx, col_name in enumerate(processed_data.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, row_data in enumerate(processed_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply formatting
        format_excel_sheet(worksheet, len(processed_data), len(processed_data.columns), processed_data)
        
        # Save the workbook
        workbook.save(build_file_path)
        workbook.close()
        
        print(f"\nSuccessfully wrote {len(processed_data)} rows to 'J.1 Automated' sheet")
        print(f"Build file saved at: {build_file_path}")
        return True
        
    except Exception as e:
        print(f"Error writing to build file: {str(e)}")
        return False


# ============================================================
# SECTION 2: CATALOG SHEET AUTOMATION FUNCTIONS
# ============================================================

def extract_to_period_number(to_period_str):
    """
    Extract numeric value from TO Period string.
    Example: "OPT PD 5" -> "5"
    """
    if not to_period_str or pd.isna(to_period_str):
        return ""
    
    # Extract numbers from the string
    match = re.search(r'\d+', str(to_period_str))
    return match.group(0) if match else ""


def format_pricing_element(pricing_element):
    """
    Format pricing element with leading zeros (e.g., 1 -> 01, 2 -> 02)
    """
    if pd.isna(pricing_element) or pricing_element == "":
        return ""
    
    try:
        # Convert to int then format with leading zero
        num = int(float(pricing_element))
        return f"{num:02d}"
    except (ValueError, TypeError):
        return str(pricing_element)


def get_clin_pricing_method_mapping(build_file_path: Path):
    """
    Read CLIN Table sheet and create mapping: EIS CLIN -> Pricing Method
    """
    try:
        # Find CLIN Table sheet (case-insensitive)
        wb = load_workbook(build_file_path, data_only=True)
        clin_sheet_name = None
        for sheet in wb.sheetnames:
            if 'clin table' in sheet.lower():
                clin_sheet_name = sheet
                break
        
        if not clin_sheet_name:
            print("Warning: CLIN Table sheet not found")
            return {}
        
        # Read as DataFrame
        clin_df = read_excel_auto(build_file_path, sheet_name=clin_sheet_name)
        clin_df.columns = clin_df.columns.str.strip()
        
        # Create mapping: Clin -> Pricing Method
        mapping = {}
        if 'Clin' in clin_df.columns and 'Pricing Method' in clin_df.columns:
            for idx, row in clin_df.iterrows():
                clin = str(row['Clin']).strip() if pd.notna(row['Clin']) else ""
                pricing_method = str(row['Pricing Method']).strip() if pd.notna(row['Pricing Method']) else ""
                if clin:
                    mapping[clin] = pricing_method
            
            print(f"Loaded {len(mapping)} CLIN -> Pricing Method mappings")
        else:
            print("Warning: Required columns not found in CLIN Table")
        
        wb.close()
        return mapping
        
    except Exception as e:
        print(f"Error reading CLIN Table: {e}")
        return {}


def get_device_class_discount_mapping(build_file_path: Path):
    """
    Read Device Class ID Table and create mapping: Device Class ID -> HHS Percentage Discount from OLP
    Handles both decimal format (0.495) and percentage format (49.5%) or percentage string ("49.5%")
    """
    try:
        # Find Device Class ID Table sheet (case-insensitive)
        wb = load_workbook(build_file_path, data_only=True)
        device_sheet_name = None
        
        for sheet in wb.sheetnames:
            if 'device class' in sheet.lower():
                device_sheet_name = sheet
                break
        
        if not device_sheet_name:
            print("Warning: Device Class ID Table sheet not found")
            return {}
        
        # Read as DataFrame
        device_df = read_excel_auto(build_file_path, sheet_name=device_sheet_name)
        device_df.columns = device_df.columns.str.strip()
        
        # Create mapping: Device Class ID -> HHS Percentage Discount from OLP
        mapping = {}
        if 'Device Class ID' in device_df.columns and 'HHS Percentage Discount from OLP' in device_df.columns:
            for idx, row in device_df.iterrows():
                device_id = str(row['Device Class ID']).strip() if pd.notna(row['Device Class ID']) else ""
                discount_raw = row['HHS Percentage Discount from OLP']
                
                # Handle different formats
                if pd.notna(discount_raw) and discount_raw != "":
                    discount_str = str(discount_raw).strip()
                    
                    # Check if it already has % sign
                    if '%' in discount_str:
                        # Already formatted as percentage (e.g., "49.5%")
                        discount_final = discount_str
                    else:
                        # Could be decimal (0.495) or number without % (49.5)
                        try:
                            discount_num = float(discount_raw)
                            
                            # If it's between 0 and 1, it's likely a decimal (0.495)
                            if 0 <= discount_num <= 1:
                                discount_percent = discount_num * 100
                                discount_final = f"{discount_percent:.2f}%"
                            else:
                                # If it's greater than 1, it's likely already a percentage number (49.5)
                                discount_final = f"{discount_num:.2f}%"
                        except (ValueError, TypeError):
                            # If conversion fails, keep as is
                            discount_final = discount_str
                else:
                    discount_final = ""
                
                if device_id:
                    mapping[device_id] = discount_final
            
            print(f"Loaded {len(mapping)} Device Class ID -> Discount mappings")
            
            # Show first 5 entries for verification
            print(f"\nFirst 5 entries in device_discount_mapping:")
            for i, (key, value) in enumerate(mapping.items()):
                if i >= 5:
                    break
                print(f"  '{key}': '{value}'")
        else:
            print("Warning: Required columns not found in Device Class ID Table")
        
        wb.close()
        return mapping
        
    except Exception as e:
        print(f"Error reading Device Class ID Table: {e}")
        return {}


def clean_device_class_id(device_id):
    """
    Clean Device Class ID to remove decimal points from float values.
    Example: 6012.0 -> "6012", "ABC123" -> "ABC123"
    """
    if pd.isna(device_id) or device_id == "":
        return ""
    
    try:
        # Try to convert to float then to int to remove decimal
        device_id_float = float(device_id)
        # Check if it's a whole number (6012.0 == 6012)
        if device_id_float.is_integer():
            return str(int(device_id_float))
        else:
            # If it has decimals, keep them
            return str(device_id_float)
    except (ValueError, TypeError):
        # If it's not a number, return as string
        return str(device_id).strip()


def get_clin_wdm_mappings(build_file_path: Path):
    """
    Read CLIN Table sheet and create mappings for wdm ICB and wdm NSP.
    Returns two dictionaries:
    - clin_icb_mapping: EIS CLIN -> "T" or "F" (from "Is ICB" column)
    - clin_nsp_mapping: EIS CLIN -> "T" or "F" (from "Not Separately Priced" column)
    """
    try:
        # Find CLIN Table sheet (case-insensitive)
        wb = load_workbook(build_file_path, data_only=True)
        clin_sheet_name = None
        for sheet in wb.sheetnames:
            if 'clin table' in sheet.lower():
                clin_sheet_name = sheet
                break
        
        if not clin_sheet_name:
            print("Warning: CLIN Table sheet not found")
            return {}, {}
        
        # Read as DataFrame
        clin_df = read_excel_auto(build_file_path, sheet_name=clin_sheet_name)
        clin_df.columns = clin_df.columns.str.strip()
        
        icb_mapping = {}
        nsp_mapping = {}
        
        # Check if required columns exist
        has_clin = 'Clin' in clin_df.columns
        has_icb = 'Is Icb' in clin_df.columns
        has_nsp = 'Not Separately Priced' in clin_df.columns
        
        if not has_clin:
            print("Warning: 'Clin' column not found in CLIN Table")
            wb.close()
            return {}, {}
        
        # Create mappings
        for idx, row in clin_df.iterrows():
            clin = str(row['Clin']).strip() if pd.notna(row['Clin']) else ""
            
            if not clin:
                continue
            
            # Map Is ICB column
            if has_icb:
                icb_value = str(row['Is Icb']).strip().upper() if pd.notna(row['Is Icb']) else ""
                if icb_value == "TRUE":
                    icb_mapping[clin] = "t"
                elif icb_value == "FALSE":
                    icb_mapping[clin] = "f"
                else:
                    icb_mapping[clin] = ""
            
            # Map Not Separately Priced column
            if has_nsp:
                nsp_value = str(row['Not Separately Priced']).strip().upper() if pd.notna(row['Not Separately Priced']) else ""
                if nsp_value == "TRUE":
                    nsp_mapping[clin] = "t"
                elif nsp_value == "FALSE":
                    nsp_mapping[clin] = "f"
                else:
                    nsp_mapping[clin] = ""
        
        print(f"Loaded {len(icb_mapping)} CLIN -> wdm ICB mappings")
        print(f"Loaded {len(nsp_mapping)} CLIN -> wdm NSP mappings")
        
        wb.close()
        return icb_mapping, nsp_mapping
        
    except Exception as e:
        print(f"Error reading CLIN Table for wdm mappings: {e}")
        return {}, {}


def apply_pricing_method_cleanup(catalog_df):
    """
    Apply cleanup rules based on YY_Price_Method column.
    Clears specific columns based on the pricing method value.
    
    Cleanup Rules:
    - ICB: Clear Orig NSC, Term NSC, Orig CJID, Term CJID
    - NONE: Clear Orig NSC, Term NSC, Orig CJID, Term CJID, Case Number, Verizon Case Description, SRE Pricing Element
    - ORIG JUR: Clear Orig NSC, Term NSC, Term CJID, Case Number, Verizon Case Description, SRE Pricing Element
    - ORIG JUR -- TERM JUR: Clear Orig NSC, Term NSC, Case Number, Verizon Case Description, SRE Pricing Element
    - ORIG NSC: Clear Term NSC, Orig CJID, Term CJID, Case Number, Verizon Case Description, SRE Pricing Element
    - ORIG NSC -- TERM NSC: Clear Orig CJID, Term CJID, Case Number, Verizon Case Description, SRE Pricing Element
    - TERM JUR: Clear Orig NSC, Term NSC, Orig CJID, Case Number, Verizon Case Description, SRE Pricing Element
    
    Additional Rule:
    - Only SRE Pricing Element 01 should have a Device Class ID; clear for all others
    """
    
    # Define cleanup rules for each pricing method
    cleanup_rules = {
        'ICB': ['Orig NSC', 'Term NSC', 'Orig CJID', 'Term CJID'],
        'NONE': ['Orig NSC', 'Term NSC', 'Orig CJID', 'Term CJID', 'Case Number', 'Verizon Case Description', 'SRE Pricing Element'],
        'ORIG JUR': ['Orig NSC', 'Term NSC', 'Term CJID', 'Case Number', 'Verizon Case Description', 'SRE Pricing Element'],
        'ORIG JUR-TERM JUR': ['Orig NSC', 'Term NSC', 'Case Number', 'Verizon Case Description', 'SRE Pricing Element'],
        'ORIG NSC': ['Term NSC', 'Orig CJID', 'Term CJID', 'Case Number', 'Verizon Case Description', 'SRE Pricing Element'],
        'ORIG NSC-TERM NSC': ['Orig CJID', 'Term CJID', 'Case Number', 'Verizon Case Description', 'SRE Pricing Element'],
        'TERM JUR': ['Orig NSC', 'Term NSC', 'Orig CJID', 'Case Number', 'Verizon Case Description', 'SRE Pricing Element']
    }
    
    print(f"\nApplying Pricing Method cleanup rules...")
    
    # Apply cleanup rules based on YY_Price_Method
    for pricing_method, columns_to_clear in cleanup_rules.items():
        # Find rows with this pricing method
        mask = catalog_df['YY_Price_Method'].apply(normalize_pricing_method) == pricing_method
        rows_affected = mask.sum()
        
        if rows_affected > 0:
            # Clear the specified columns for matching rows
            for col in columns_to_clear:
                if col in catalog_df.columns:
                    catalog_df.loc[mask, col] = ""
            
            print(f"  {pricing_method}: Cleared {len(columns_to_clear)} columns for {rows_affected} rows")
    
    # Additional rule: Only SRE Pricing Element 01 should have Device Class ID
    # Clear Device Class ID for all rows where SRE Pricing Element is NOT "01"
    non_01_mask = catalog_df['SRE Pricing Element'].astype(str).str.strip() != "01"
    device_class_cleared = non_01_mask.sum()
    
    if device_class_cleared > 0:
        catalog_df.loc[non_01_mask, 'SRE Device Class ID'] = ""
        catalog_df.loc[non_01_mask, 'HHS Discount from OLP'] = ""  # Also clear the discount since it's based on Device Class ID
        print(f"  Cleared Device Class ID for {device_class_cleared} rows (SRE Pricing Element != '01')")
    
    print(f"Pricing Method cleanup complete.\n")
    
    return catalog_df


def deduplicate_catalog(catalog_df):
    """
    Remove duplicate rows from catalog based on pricing method-specific lookup values.
    Keeps only the first occurrence of each unique lookup value.
    
    Special handling for Price Request Number:
    - If duplicate has same Price Request Number as original: delete duplicate
    - If duplicate has different Price Request Number: concatenate with original (comma-separated), then delete duplicate
    
    Lookup value composition by pricing method:
    - ICB: EIS CLIN/Case Number/SRE Pricing Element/TO Period
    - Orig NSC: EIS CLIN/Orig NSC/SRE Pricing Element/TO Period
    - Orig NSC - Term NSC: EIS CLIN/Orig NSC/Term NSC/SRE Pricing Element/TO Period
    - Term NSC: EIS CLIN/Term NSC/SRE Pricing Element/TO Period
    - Orig JUR (Orig CJID): EIS CLIN/Orig CJID/SRE Pricing Element/TO Period
    - Orig JUR - Term JUR (Orig CJID - Term CJID): EIS CLIN/Orig CJID/Term CJID/SRE Pricing Element/TO Period
    - Term JUR (Term CJID): EIS CLIN/Term CJID/SRE Pricing Element/TO Period
    """
    
    print(f"\nApplying deduplication logic...")
    
    # Define lookup value logic for each pricing method
    pricing_method_lookup_map = {
        'ICB': ['EIS CLIN', 'Case Number', 'SRE Pricing Element','TO Period'],
        'ORIG NSC': ['EIS CLIN', 'Orig NSC', 'SRE Pricing Element','TO Period'],
        'ORIG NSC-TERM NSC': ['EIS CLIN', 'Orig NSC', 'Term NSC', 'SRE Pricing Element','TO Period'],
        'TERM NSC': ['EIS CLIN', 'Term NSC', 'SRE Pricing Element','TO Period'],
        'ORIG JUR': ['EIS CLIN', 'Orig CJID', 'SRE Pricing Element','TO Period'],
        'ORIG JUR-TERM JUR': ['EIS CLIN', 'Orig CJID', 'Term CJID', 'SRE Pricing Element','TO Period'],
        'TERM JUR': ['EIS CLIN', 'Term CJID', 'SRE Pricing Element','TO Period'],
    }
    
    # Create lookup value column
    def create_lookup_value(row):
        """
        Create lookup value by concatenating relevant columns based on pricing method.
        """
        pricing_method = normalize_pricing_method(row['YY_Price_Method'])
        
        # Get the columns to concatenate for this pricing method.
        # For recognized pricing methods, use the business-specific lookup key.
        # For blank/unknown methods, use a conservative full-row identity so
        # manual-review rows stay visible, but do not include Price Request Number
        # in the key. Duplicate Catalog items across PRs should collapse into one
        # row with the PR numbers concatenated.
        columns_to_use = pricing_method_lookup_map.get(pricing_method)
        if not columns_to_use:
            columns_to_use = [
                'EIS CLIN',
                'EIS CLIN Name',
                'Orig NSC',
                'Term NSC',
                'Orig CJID',
                'Term CJID',
                'Case Number',
                'SRE Pricing Element',
                'TO Period',
                'HHS Price',
            ]
        
        # Concatenate values with "/" separator, converting to string and stripping whitespace
        lookup_parts = []
        for col in columns_to_use:
            if col in row.index:
                value = str(row[col]).strip() if pd.notna(row[col]) and row[col] != "" else ""
            else:
                value = ""
            lookup_parts.append(value)
        
        return "/".join(lookup_parts)
    
    # Add lookup value column
    catalog_df['_lookup_value'] = catalog_df.apply(create_lookup_value, axis=1)
    
    # Process duplicates with special Price Request Number handling
    grouped = catalog_df.groupby('_lookup_value')
    
    rows_to_remove = []
    pr_concatenations = 0
    duplicate_details = []
    
    for lookup_val, group in grouped:
        if len(group) > 1:  # Found duplicates
            # Get the first occurrence (will be kept)
            first_idx = group.index[0]
            first_pr = str(catalog_df.at[first_idx, 'Price Request Number']).strip() if pd.notna(catalog_df.at[first_idx, 'Price Request Number']) else ""
            
            # Collect all unique Price Request Numbers from duplicates
            unique_prs = set()
            if first_pr and first_pr != "":
                unique_prs.add(first_pr)
            
            # Collect duplicate info
            dup_prs = []
            
            # Process each duplicate
            for dup_idx in group.index[1:]:
                dup_pr = str(catalog_df.at[dup_idx, 'Price Request Number']).strip() if pd.notna(catalog_df.at[dup_idx, 'Price Request Number']) else ""
                dup_prs.append(dup_pr if dup_pr else "(empty)")
                
                # Add to unique PRs if different and not empty
                if dup_pr and dup_pr != "" and dup_pr not in unique_prs:
                    unique_prs.add(dup_pr)
                
                # Mark duplicate for removal
                rows_to_remove.append(dup_idx)
            
            # Store duplicate details for reporting
            duplicate_details.append({
                'lookup': lookup_val,
                'kept_row': first_idx,
                'kept_pr': first_pr if first_pr else "(empty)",
                'removed_rows': group.index[1:].tolist(),
                'removed_prs': dup_prs,
                'count': len(group)
            })
            
            # If we have multiple unique PRs, concatenate them
            if len(unique_prs) > 1:
                concatenated_pr = "/".join(sorted(unique_prs))
                pr_concatenations += 1
                # Update the Price Request Number in the kept row
                catalog_df.at[first_idx, 'Price Request Number'] = concatenated_pr
    
    # Print duplicate details
    if duplicate_details:
        print(f"\nDuplicate elements found:")
        print(f"{'='*80}")
        for detail in duplicate_details:
            print(f"\nLookup Value: {detail['lookup']}")
            print(f"  Total occurrences: {detail['count']}")
            print(f"  Kept row {detail['kept_row']} (PR: {detail['kept_pr']})")
            print(f"  Removed rows: {detail['removed_rows']} (PRs: {detail['removed_prs']})")
        print(f"{'='*80}\n")
    
    # Remove duplicates
    duplicates_mask = catalog_df.index.isin(rows_to_remove)
    catalog_df_deduped = catalog_df[~duplicates_mask].copy()
    
    # Drop the temporary lookup value column
    catalog_df_deduped = catalog_df_deduped.drop(columns=['_lookup_value'])
    
    # Summary
    rows_removed = len(catalog_df) - len(catalog_df_deduped)
    print(f"  Rows before: {len(catalog_df)}")
    print(f"  Rows after: {len(catalog_df_deduped)}")
    print(f"  Duplicates removed: {rows_removed}")
    if pr_concatenations > 0:
        print(f"  Price Request Numbers concatenated: {pr_concatenations}")
    print(f"Deduplication complete.\n")
    
    return catalog_df_deduped


def create_catalog_sheet(build_file_path: Path):
    """
    Creates or updates the Catalog worksheet in the same build file.
    If Catalog sheet exists, it will be replaced with the new automated version.
    """
    
    print(f"\n{'='*60}")
    print("CREATING CATALOG SHEET")
    print(f"{'='*60}")
    print(f"Reading {build_file_path.name}...")
    
    # Load the original workbook
    wb_original = load_workbook(build_file_path)
    
    # Check if J.1 Automated sheet exists (use it instead of J.1 if available)
    j1_sheet_name = None
    for sheet in wb_original.sheetnames:
        if 'j.1 automated' in sheet.lower():
            j1_sheet_name = sheet
            break
    
    # If J.1 Automated not found, look for J.1
    if not j1_sheet_name:
        for sheet in wb_original.sheetnames:
            if 'j.1' in sheet.lower() or 'j1' in sheet.lower():
                j1_sheet_name = sheet
                break
    
    if not j1_sheet_name:
        print("Error: J.1 or J.1 Automated worksheet not found!")
        wb_original.close()
        return None
    
    print(f"Found J.1 sheet: {j1_sheet_name}")
    
    # Read J.1 sheet as DataFrame
    j1_df = read_excel_auto(build_file_path, sheet_name=j1_sheet_name)
    j1_df.columns = j1_df.columns.str.strip()
    
    print(f"J.1 sheet has {len(j1_df)} rows")
    
    # Get lookup mappings from other sheets
    clin_pricing_mapping = get_clin_pricing_method_mapping(build_file_path)
    device_discount_mapping = get_device_class_discount_mapping(build_file_path)
    clin_icb_mapping, clin_nsp_mapping = get_clin_wdm_mappings(build_file_path)

    # Define direct column mappings: J.1 column -> Catalog column
    direct_mappings = {
        'EIS CLIN': 'EIS CLIN',
        'EIS CLIN Name': 'EIS CLIN Name',
        'Orig NSC': 'Orig NSC',
        'Term NSC': 'Term NSC',
        'Orig CJID': 'Orig CJID',
        'Term CJID': 'Term CJID',
        'Band Low': 'Band Low',
        'Band High': 'Band High',
        'Case Number': 'Case Number',
        'Catalog Cse ID': 'SRE Device Class ID',
        'HHS Price': 'HHS Price',
        'EIS Frequency': 'EIS Frequency',
        'Charging Unit': 'Charging Unit',
        'Price Start Date': 'Price Start Date',
        'Price End Date': 'Price End Date',
        'TO Period': 'TO Period',
        'ICB Case Description': 'Verizon Case Description',
        'Pricing Request #': 'Price Request Number',
        'SRE Pricing Element': 'SRE Pricing Element',
        'MOD Request  Description': 'MOD Request Description',
        'Revision Description': 'Revision Description',
        'GSA  Awarded EIS Price': 'GSA Awarded Price',
    }
    
    # Create catalog dataframe with direct mappings
    catalog_data = {}
    
    for j1_col, catalog_col in direct_mappings.items():
        if j1_col in j1_df.columns:
            if catalog_col == 'TO Period':
                # Special handling: extract numeric value only
                catalog_data[catalog_col] = j1_df[j1_col].apply(extract_to_period_number)
            elif catalog_col == 'SRE Pricing Element':
                # Special handling: format with leading zeros (01, 02, etc.)
                catalog_data[catalog_col] = j1_df[j1_col].apply(format_pricing_element)
            elif catalog_col == 'SRE Device Class ID':
                # Special handling: remove decimal point from floats (6012.0 -> 6012)
                catalog_data[catalog_col] = j1_df[j1_col].apply(clean_device_class_id)
            else:
                catalog_data[catalog_col] = j1_df[j1_col]
        else:
            print(f"Warning: Column '{j1_col}' not found in J.1 sheet")
            catalog_data[catalog_col] = ""
    
    # Define all Catalog columns in the correct order
    catalog_columns = [
        'YY_Price_Method',
        'EIS CLIN',
        'EIS CLIN Name',
        'EIS Frequency',
        'Charging Unit',
        'Orig NSC',
        'Term NSC',
        'Orig CJID',
        'Term CJID',
        'HHS Price',
        'wdm ICB',
        'wdm NSP',
        'Case Number',
        'Verizon Case Description',
        'SRE Pricing Element',
        'SRE Device Class ID',
        'HHS Discount from OLP',
        'Band Low',
        'Band High',
        'TO Period',
        'Price Start Date',
        'Price End Date',
        'Mod Number',
        'Price Request Number',
        'MOD Request Description',
        'Revision Description',
        'GSA Awarded Price'
    ]
    
    # Create DataFrame with all columns (empty ones will be filled later with indirect mappings)
    catalog_df = pd.DataFrame(catalog_data)
    
    # Add empty columns for those not yet mapped
    for col in catalog_columns:
        if col not in catalog_df.columns:
            catalog_df[col] = ""
    
    # Reorder columns
    catalog_df = catalog_df[catalog_columns]
    
    # Populate YY_Price_Method using CLIN lookup
    if 'EIS CLIN' in catalog_df.columns and clin_pricing_mapping:
        catalog_df['YY_Price_Method'] = catalog_df['EIS CLIN'].apply(
            lambda x: clin_pricing_mapping.get(str(x).strip(), "") if pd.notna(x) else ""
        )
        print(f"Populated YY_Price_Method for {catalog_df['YY_Price_Method'].notna().sum()} rows")
    
    # Populate HHS Discount from OLP using Device Class ID lookup
    if 'SRE Device Class ID' in catalog_df.columns and device_discount_mapping:
        catalog_df['HHS Discount from OLP'] = catalog_df['SRE Device Class ID'].apply(
            lambda x: device_discount_mapping.get(str(x).strip(), "") if pd.notna(x) else ""
        )
        
        successful_matches = (catalog_df['HHS Discount from OLP'] != "").sum()
        print(f"Populated HHS Discount from OLP for {successful_matches}/{len(catalog_df)} rows")

    # Populate wdm ICB using CLIN lookup
    if 'EIS CLIN' in catalog_df.columns and clin_icb_mapping:
        catalog_df['wdm ICB'] = catalog_df['EIS CLIN'].apply(
            lambda x: clin_icb_mapping.get(str(x).strip(), "") if pd.notna(x) else ""
        )
        
        successful_icb = (catalog_df['wdm ICB'] != "").sum()
        print(f"Populated wdm ICB for {successful_icb}/{len(catalog_df)} rows")
    
    # Populate wdm NSP using CLIN lookup
    if 'EIS CLIN' in catalog_df.columns and clin_nsp_mapping:
        catalog_df['wdm NSP'] = catalog_df['EIS CLIN'].apply(
            lambda x: clin_nsp_mapping.get(str(x).strip(), "") if pd.notna(x) else ""
        )
        
        successful_nsp = (catalog_df['wdm NSP'] != "").sum()
        print(f"Populated wdm NSP for {successful_nsp}/{len(catalog_df)} rows")

    # Apply deduplication before display cleanup so lookup keys are built
    # from the full J.1-derived values, not from fields that cleanup may blank
    # for Catalog presentation.
    catalog_df = deduplicate_catalog(catalog_df)

    # Apply Catalog display cleanup after deduplication.
    catalog_df = apply_pricing_method_cleanup(catalog_df)
    
    # ============================================================
    # Write Catalog sheet back to the same file
    # ============================================================
    
    # Remove existing Catalog sheet if it exists
    if 'Catalog' in wb_original.sheetnames:
        del wb_original['Catalog']
        print("\nRemoved existing Catalog sheet")
    
    # Create new Catalog sheet
    ws_catalog = wb_original.create_sheet('Catalog')
    
    # Write headers with BLACK fill and WHITE font
    for col_idx, col_name in enumerate(catalog_columns, start=1):
        cell = ws_catalog.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write data
    for row_idx, row_data in enumerate(catalog_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_catalog.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Auto-adjust column widths
    for col in ws_catalog.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_catalog.column_dimensions[column].width = max(adjusted_width, 12)
    
    # Freeze first row
    ws_catalog.freeze_panes = "A2"
    
    # Add auto-filter to all columns
    ws_catalog.auto_filter.ref = f"A1:{ws_catalog.cell(row=1, column=len(catalog_columns)).coordinate}"
    
    # Save the workbook back to the original file
    wb_original.save(build_file_path)
    wb_original.close()
    
    print(f"\nCatalog sheet created/updated in: {build_file_path}")
    print(f"  Contains: {len(catalog_df)} rows with automated Catalog data (with filters)")
    
    return build_file_path


# ============================================================
# MAIN EXECUTION
# ============================================================

def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Create the build workbook from selected build template, coversheets, and PR files."
    )
    parser.add_argument("--build-file", help="Build template/source workbook to copy and update")
    parser.add_argument("--coversheets-dir", help="Folder containing approved coversheet workbooks")
    parser.add_argument("--pr-dir", help="Folder containing PR workbooks")
    parser.add_argument("--output-file", help="Output build workbook path")
    return parser.parse_args(argv)


def configure_paths(args):
    global build_file_input, build_file_output, PR_DIR, COVERSHEET_FILES_DIR
    build_file_input = choose_existing_file(args.build_file, "Select the build template/source workbook", EXCEL_FILETYPES)
    COVERSHEET_FILES_DIR = choose_existing_dir(args.coversheets_dir, "Select the folder containing approved coversheets")
    PR_DIR = choose_existing_dir(args.pr_dir, "Select the folder containing PR workbooks")
    build_file_output = choose_save_file(args.output_file, "Save the generated build workbook as", "build_file.xlsx", EXCEL_FILETYPES)


def main(argv=None):
    """Main function to run both J.1 and Catalog automation."""
    args = parse_args(argv)
    configure_paths(args)
    build_file_output.parent.mkdir(parents=True, exist_ok=True)

    # Copy the selected pristine build file to the selected output path before modifying.
    if not build_file_input.exists():
        print(f"Error: Build file not found at {build_file_input}")
        return {'j1_rows': 0, 'catalog_rows': 0, 'success': False}

    shutil.copy2(build_file_input, build_file_output)
    print(f"Copied build file to: {build_file_output}")

    print("=" * 60)
    print("COMBINED AUTOMATION SCRIPT")
    print("=" * 60)
    print(f"PR Path: {PR_DIR}")
    print(f"Coversheet Path: {COVERSHEET_FILES_DIR}")
    print(f"Build File Path: {build_file_output}")
    print("=" * 60)

    # ============================================================
    # STEP 1: J.1 SHEET AUTOMATION
    # ============================================================
    print("\n" + "=" * 60)
    print("STEP 1: PROCESSING PR AND COVERSHEET FILES")
    print("=" * 60)

    processed_data = process_all_coversheets(
        COVERSHEET_FILES_DIR, PR_DIR, build_file_output
    )

    j1_success = write_to_build_file(processed_data, build_file_output)

    if not j1_success:
        print("\nError: J.1 automation failed. Stopping execution.")
        return {'j1_rows': 0, 'catalog_rows': 0, 'success': False}

    # ============================================================
    # STEP 2: CATALOG SHEET AUTOMATION
    # ============================================================
    print("\n" + "=" * 60)
    print("STEP 2: CREATING CATALOG SHEET FROM J.1 DATA")
    print("=" * 60)

    catalog_result = create_catalog_sheet(build_file_output)

    catalog_success = catalog_result is not None

    # ============================================================
    # FINAL SUMMARY
    # ============================================================
    print("\n" + "=" * 60)
    print("AUTOMATION COMPLETE - FINAL SUMMARY")
    print("=" * 60)
    j1_row_count = len(processed_data) if not processed_data.empty else 0
    print(f"J.1 Automated sheet: {j1_row_count} rows")
    print("Catalog sheet: Created successfully" if catalog_success
          else "Catalog sheet: Failed")
    print(f"Build file location: {build_file_output}")
    status = ('SUCCESS' if (j1_success and catalog_success)
              else 'PARTIAL SUCCESS' if j1_success else 'FAILED')
    print(f"Overall Status: {status}")
    print("=" * 60)

    return {
        'j1_rows': j1_row_count,
        'catalog_success': catalog_success,
        'build_file_location': str(build_file_output),
        'success': j1_success and catalog_success
    }


if __name__ == "__main__":
    main()