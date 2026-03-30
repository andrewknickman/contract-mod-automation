import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import os
from typing import Dict, List, Any, Optional

from workflow_io import choose_directory, choose_file, choose_save_file



# ─── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR = None
INPUT_DIR = None
OUTPUT_DIR = None
COVERSHEET_FILES_DIR = None

clin_table_file = None
overview_output_file = None


def cell_is_checked(val) -> bool:
    """
    Return True if a cell value represents a checked checkbox (linked TRUE).
    Handles Excel-linked booleans and a few common textual markers.
    """
    if val is True:
        return True
    if isinstance(val, (int, float)) and val == 1:
        return True
    if isinstance(val, str):
        v = val.strip().lower()
        return v in {"true", "yes", "y", "x", "✓", "checked", "1"}
    return False


def find_checked_checkbox(sheet, row: int, col_range: range) -> Optional[str]:
    """
    Find which checkbox is checked in a given row and column range.
    
    Args:
        sheet: The worksheet object
        row: Row number
        col_range: Range of columns to check
        
    Returns:
        The value of the cell where checkbox is checked, or None
    """
    for col in col_range:
        cell = sheet.cell(row=row, column=col)
        if cell_is_checked(cell.value):
            # Try to read the label from the cell above (common layout),
            # otherwise from the same cell if it contains text.
            label_cell = sheet.cell(row=row - 1, column=col) if row > 1 else cell
            if label_cell.value and str(label_cell.value).strip():
                return str(label_cell.value).strip()
            
            header_cell = sheet.cell(row=row, column=col)
            if header_cell.value and not cell_is_checked(header_cell.value):
                return str(header_cell.value).strip()
    
    # Fallback: if any cell in the range has a non-empty non-falsey text, return it
    for col in col_range:
        v = sheet.cell(row=row, column=col).value
        if v and str(v).strip() and str(v).strip().lower() not in {"false", "no", "0"}:
            return str(v).strip()
    
    return None


def get_opdiv_value(sheet) -> str:
    """
    Extract OpDiv value from checkboxes in row 12, columns C-K via linked cells.
    (Kept for compatibility, but not used when xlwings-only is requested.)
    """
    opdiv_map = {
        3: 'CDC',    # C12
        4: 'CMS',    # D12
        5: 'FDA',    # E12
        6: 'HRSA',   # F12
        7: 'IHS',    # G12
        8: 'NIH',    # H12
        9: 'OIG',    # I12
        10: 'OS',    # J12
        11: 'ENMOD'  # K12
    }
    for col, opdiv_name in opdiv_map.items():
        cell = sheet.cell(row=13, column=col)
        if cell_is_checked(cell.value):
            return opdiv_name
    return ""


def get_ticket_value(sheet) -> str:
    """
    Extract ticket value from cells K7, L7, M7.
    """
    ticket_parts = []
    for col in [11, 12, 13]:  # K, L, M columns
        cell = sheet.cell(row=7, column=col)
        if cell.value and str(cell.value).strip():
            ticket_parts.append(str(cell.value).strip())
    return ' '.join(ticket_parts)


def get_type_of_mod(sheet, checked_row_hint: Optional[int] = None, use_hint_only: bool = False) -> tuple:
    """
    Extract Type of Mod and its description.

    Args:
        sheet: Mod Description worksheet
        checked_row_hint: optional row number (15/20/25/30) if already known via shapes/xlwings
        use_hint_only: if True, do NOT infer from cells; only use the hint

    Returns:
        Tuple of (type_of_mod, description)
    """
    mod_rows = [15, 20, 25, 30]
    type_of_mod = ""
    description = ""
    checked_row = checked_row_hint

    if use_hint_only and not checked_row:
        return "", ""

    # If we didn't force hint-only, allow linked-cell detection
    if not use_hint_only and not checked_row:
        for row in mod_rows:
            cell_a = sheet.cell(row=row, column=1)
            if cell_is_checked(cell_a.value):
                checked_row = row
                break

    if checked_row:
        mod_text_parts = []
        for col in range(2, 15):  # B(2) to N(14)
            cell = sheet.cell(row=checked_row, column=col)
            if cell.value and str(cell.value).strip():
                mod_text_parts.append(str(cell.value).strip())
        type_of_mod = ' '.join(mod_text_parts)

        if checked_row == 15:
            description_rows = range(16, 19)
        elif checked_row == 20:
            description_rows = range(21, 24)
        elif checked_row == 25:
            description_rows = range(26, 29)
        elif checked_row == 30:
            description_rows = range(31, 34)
        else:
            description_rows = []

        description_parts = []
        for desc_row in description_rows:
            row_text = []
            for col in range(1, 15):  # A..N
                cell = sheet.cell(row=desc_row, column=col)
                if cell.value and str(cell.value).strip():
                    row_text.append(str(cell.value).strip())
            if row_text:
                description_parts.append(' '.join(row_text))
        description = ' '.join(description_parts)

    # If use_hint_only and hint led to empty text, just return empty strings.
    if use_hint_only:
        return type_of_mod, description

    # Fallback by content only if not enforcing hint-only
    if not type_of_mod and not use_hint_only:
        for row in mod_rows:
            row_text_parts = []
            for col in range(2, 15):
                cell = sheet.cell(row=row, column=col)
                if cell.value and str(cell.value).strip():
                    row_text_parts.append(str(cell.value).strip())
            if row_text_parts and len(' '.join(row_text_parts)) > 10:
                type_of_mod = ' '.join(row_text_parts)
                if row == 15:
                    description_rows = range(16, 19)
                elif row == 20:
                    description_rows = range(21, 24)
                elif row == 25:
                    description_rows = range(26, 29)
                elif row == 30:
                    description_rows = range(31, 34)
                else:
                    description_rows = []
                description_parts = []
                for desc_row in description_rows:
                    row_text = []
                    for col in range(1, 15):
                        cell = sheet.cell(row=desc_row, column=col)
                        if cell.value and str(cell.value).strip():
                            row_text.append(str(cell.value).strip())
                    if row_text:
                        description_parts.append(' '.join(row_text))
                description = ' '.join(description_parts)
                break

    return type_of_mod, description


def get_funding_info(sheet) -> str:
    """
    Extract funding information from rows 60-63, columns A-N.
    """
    funding_parts = []
    for row in range(60, 64):
        row_text = []
        for col in range(1, 15):
            cell = sheet.cell(row=row, column=col)
            if cell.value and str(cell.value).strip():
                row_text.append(str(cell.value).strip())
        if row_text:
            funding_parts.append(' '.join(row_text))
    return ' '.join(funding_parts)


def get_standard_notes(sheet) -> str:
    """
    Extract standard notes from cells D7 to F7.
    """
    notes_parts = []
    for col in range(4, 7):  # D..F
        cell = sheet.cell(row=7, column=col)
        if cell.value and str(cell.value).strip():
            notes_parts.append(str(cell.value).strip())
    return ' '.join(notes_parts)


def load_clin_lookup_table(lookup_file_path: str) -> Dict[str, str]:
    """
    Load the CLIN Table lookup file and create a mapping of CLIN to Service Id.
    Maps from 'Clin' column to 'Service Id' column in the lookup file.
    Supports both .xls and .xlsx formats via pandas.
    """
    try:
        df = pd.read_excel(lookup_file_path)
        df.columns = df.columns.str.strip()

        clin_col = 'Clin' if 'Clin' in df.columns else ('CLIN' if 'CLIN' in df.columns else None)
        service_col = 'Service Id' if 'Service Id' in df.columns else None
        if not clin_col or not service_col:
            return {}

        clin_to_service = {}
        for _, row in df.iterrows():
            clin = row[clin_col]
            service_id = row[service_col]
            if pd.notna(clin) and pd.notna(service_id):
                clin_str = str(clin).strip().upper()
                service_str = str(service_id).strip()
                clin_to_service[clin_str] = service_str
                if len(clin_str) > 2 and clin_str[:2].isalpha():
                    prefix = clin_str[:2]
                    try:
                        number_part = clin_str[2:]
                        if number_part.isdigit():
                            number_without_zeros = str(int(number_part))
                            clin_to_service[prefix + number_without_zeros] = service_str
                    except Exception:
                        pass
        return clin_to_service
    except Exception:
        return {}


def get_services_from_clin_tables(wb, lookup_table: Dict[str, str]) -> str:
    """
    Extract unique Service IDs from CLIN tables using the lookup table.
    """
    if not lookup_table:
        return ""
    service_ids = set()
    for sheet_name in ['CLIN Table (Current OP)', 'CLIN Table (OY)']:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = {}
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header:
                    headers[str(header).strip()] = col
            clin_col = headers.get('CLIN', None)
            if not clin_col:
                continue
            for row in range(2, sheet.max_row + 1):
                clin = sheet.cell(row=row, column=clin_col).value
                if clin:
                    clin_str = str(clin).strip().upper()
                    if clin_str in lookup_table:
                        service_ids.add(lookup_table[clin_str])
                    else:
                        if len(clin_str) > 2 and clin_str[:2].isalpha():
                            prefix = clin_str[:2]
                            try:
                                number_part = clin_str[2:]
                                if number_part.isdigit():
                                    number_without_zeros = str(int(number_part))
                                    alt_clin = prefix + number_without_zeros
                                    if alt_clin in lookup_table:
                                        service_ids.add(lookup_table[alt_clin])
                            except:
                                pass
    return ', '.join(sorted(service_ids))


def process_clin_table(sheet) -> Dict[str, Any]:
    """
    Process CLIN Table (Current OP) sheet.
    """
    results = {
        'clin_count': 0,
        'opdiv_approval': 'No',
        'fr_needed': 'No',
        'fr_count': 0,
        'msa_mod': ''
    }
    headers = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            headers[str(header).strip()] = col

    opdiv_yes_count = 0
    fr_pending_na_found = False
    fr_items = []
    msa_mods = {}

    for row in range(2, sheet.max_row + 1):
        opdiv_col = headers.get('OpDiv Approval (Yes/ No)', headers.get('OpDiv Approval', None))
        if opdiv_col:
            opdiv_value = sheet.cell(row=row, column=opdiv_col).value
            if opdiv_value and str(opdiv_value).strip().lower() in ['yes', 'y']:
                opdiv_yes_count += 1
                results['opdiv_approval'] = 'Yes'
                fr_col = headers.get('F&R Status (Pending, Approved, or N/A)', headers.get('F&R Status', None))
                if fr_col:
                    fr_value = sheet.cell(row=row, column=fr_col).value
                    if fr_value and str(fr_value).strip().lower() in ['pending', 'n/a', 'na']:
                        fr_pending_na_found = True

        fr_col = headers.get('F&R Status (Pending, Approved, or N/A)', headers.get('F&R Status', None))
        if fr_col:
            fr_value = sheet.cell(row=row, column=fr_col).value
            if fr_value and str(fr_value).strip().lower() in ['pending', 'approved']:
                fr_items.append(fr_value)

        msa_mod_col = headers.get('MSA Mod Number', headers.get('MSA Mod', None))
        msa_status_col = headers.get('MSA Status (Pending, Awarded, or N/A)', headers.get('MSA Status', None))
        if msa_mod_col and msa_status_col:
            msa_mod_num = sheet.cell(row=row, column=msa_mod_col).value
            msa_status = sheet.cell(row=row, column=msa_status_col).value
            num_str = str(msa_mod_num).strip() if msa_mod_num else ''
            status_str = str(msa_status).strip() if msa_status else ''
            if num_str:
                msa_mods[num_str] = status_str

    results['clin_count'] = opdiv_yes_count
    results['fr_needed'] = 'Yes' if fr_pending_na_found else 'No'
    results['fr_count'] = len(fr_items)

    if msa_mods:
        msa_mod_list = []
        for num, status in msa_mods.items():
            if status:
                msa_mod_list.append(f"{status} {num}")
            else:
                msa_mod_list.append(f"{num}")
        results['msa_mod'] = ', '.join(msa_mod_list)

    return results


# --------- xlwings-based helpers (no fallback) ---------
# REPLACE your _collect_checkbox_states_xlwings with this version.
# (Some Excel files don’t surface Form Controls under Shapes; use CheckBoxes())

def _collect_checkbox_states_xlwings(path: str, sheet_name: str = "Mod Description",
                                     app=None) -> Dict[tuple, Dict[str, Any]]:
    """
    Returns {(row, col): {'checked': bool, 'caption': str}} for all Form Control checkboxes.
    Uses Worksheet.CheckBoxes (preferred) and falls back to Shapes only if needed to build the map.

    If *app* is provided (an existing xlwings App instance), it will be reused
    instead of creating (and quitting) a throwaway instance.
    """
    import xlwings as xw

    owns_app = app is None
    if owns_app:
        app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(path, read_only=True)
        sht = wb.sheets[sheet_name]
        out: Dict[tuple, Dict[str, Any]] = {}

        # 1) Preferred: worksheet-level Form Controls
        try:
            for cb in sht.api.CheckBoxes():
                tl = cb.TopLeftCell
                out[(tl.Row, tl.Column)] = {
                    "checked": bool(cb.Value == 1),
                    "caption": str(cb.Caption or "").strip(),
                }
        except Exception:
            pass

        # 2) If nothing found, also include Shapes-based controls
        try:
            XL_FORM_CHECKBOX = -4146  # xlCheckBox
            for shp in sht.api.Shapes:
                try:
                    if hasattr(shp, "FormControlType") and shp.FormControlType == XL_FORM_CHECKBOX:
                        tl = shp.TopLeftCell
                        out[(tl.Row, tl.Column)] = {
                            "checked": bool(shp.ControlFormat.Value == 1),
                            "caption": str(shp.TextFrame.Characters().Text or "").strip(),
                        }
                except Exception:
                    continue
        except Exception:
            pass

        wb.close()
        return out
    finally:
        if owns_app:
            app.quit()


# UPDATE the exact-position map to row 11 (per the anchors in your workbook)

def get_opdiv_value_xlwings(path: str, app=None) -> str:
    states = _collect_checkbox_states_xlwings(path, "Mod Description", app=app)
    # OpDiv checkboxes' TopLeftCell are on row 11, columns C..K
    pos2name = {
        (11, 3): "CDC", (11, 4): "CMS", (11, 5): "FDA",
        (11, 6): "HRSA", (11, 7): "IHS", (11, 8): "NIH",
        (11, 9): "OIG", (11,10): "OS",  (11,11): "ENMOD",
    }
    for pos, name in pos2name.items():
        info = states.get(pos)
        if info and info.get("checked") is True:
            return name
    return ""


def get_type_of_mod_row_xlwings(path: str, app=None) -> Optional[int]:
    """
    Resolve Type of Mod strictly by the checkbox whose TopLeftCell
    is on column A (1) at rows 15, 20, 25, or 30. No fallbacks.
    """
    states = _collect_checkbox_states_xlwings(path, "Mod Description", app=app)

    exact_pos_to_row = {
        (15, 1): 15,  # A15
        (20, 1): 20,  # A20
        (25, 1): 25,  # A25
        (30, 1): 30,  # A30
    }

    for pos, row in exact_pos_to_row.items():
        info = states.get(pos)
        if info and info.get("checked") is True:
            return row

    return None



def process_coversheet(filepath: Path, clin_lookup_table: Dict[str, str] = None,
                       xlwings_app=None) -> Dict[str, Any]:
    """
    Process a single coversheet file and extract all required data.
    Uses xlwings exclusively for checkbox-driven fields (no fallback).

    If *xlwings_app* is provided, the shared xlwings App instance is reused
    for all checkbox reads instead of spawning a new Excel process each time.
    """
    wb = load_workbook(filepath, data_only=True)
    
    overview_columns = [
        'PR#', 'Version', 'OpDiv', 'Ticket', '# of CLINs (Line items in current OP)',
        'OpDiv Approval', 'Services', 'SF30 Description', 'F&R Needed', '# of CLINs',
        'F&R Notes', 'Type of F&R', 'Liaison', 'SME', 'Type of Mod',
        'Coversheet Description', 'Funding', '12M+ CLIN', 'MSA Mod', 'Standard Notes',
        'New Req?', 'Section C Update Needed?', 'Tara Matthews Approval', 'Cristy Scott Approval'
    ]
    result = {col: '' for col in overview_columns}
    
    mod_sheet = wb['Mod Description']
    
    # 1. PR#
    result['PR#'] = str(mod_sheet.cell(row=3, column=4).value) if mod_sheet.cell(row=3, column=4).value else ''
    # 2. Version
    result['Version'] = str(mod_sheet.cell(row=5, column=4).value) if mod_sheet.cell(row=5, column=4).value else ''
    # 3. OpDiv (xlwings only)
    result['OpDiv'] = get_opdiv_value_xlwings(str(filepath), app=xlwings_app)  # no fallback
    # 4. Ticket
    result['Ticket'] = get_ticket_value(mod_sheet)
    # 13. Liaison
    result['Liaison'] = str(mod_sheet.cell(row=3, column=11).value) if mod_sheet.cell(row=3, column=11).value else ''
    # 14. SME
    result['SME'] = str(mod_sheet.cell(row=5, column=11).value) if mod_sheet.cell(row=5, column=11).value else ''
    # 15 & 16. Type of Mod / Coversheet Description (xlwings only for the checkbox row)
    checked_row_hint = get_type_of_mod_row_xlwings(str(filepath), app=xlwings_app)
    type_of_mod, description = get_type_of_mod(mod_sheet, checked_row_hint=checked_row_hint, use_hint_only=True)
    result['Type of Mod'] = type_of_mod
    result['Coversheet Description'] = description
    # 17. Funding
    result['Funding'] = get_funding_info(mod_sheet)
    # 20. Standard Notes
    result['Standard Notes'] = get_standard_notes(mod_sheet)
    # 7. Services
    result['Services'] = get_services_from_clin_tables(wb, clin_lookup_table) if clin_lookup_table else ''
    
    # CLIN Table (Current OP)
    clin_sheet = wb['CLIN Table (Current OP)']
    clin_data = process_clin_table(clin_sheet)
    result['# of CLINs (Line items in current OP)'] = str(clin_data['clin_count'])
    result['OpDiv Approval'] = clin_data['opdiv_approval']
    result['F&R Needed'] = clin_data['fr_needed']
    result['# of CLINs'] = str(clin_data['fr_count'])
    result['MSA Mod'] = clin_data['msa_mod']
    
    wb.close()
    return result


def apply_styling(worksheet):
    """
    Apply formatting and styling to the overview worksheet.
    """
    from openpyxl.styles import Border, Side
    
    header_font = Font(name='Calibri', size=12, bold=True)
    cell_font = Font(name='Calibri', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    color_groups = [
        (range(1, 3), 'F2F2F2'),
        (range(3, 9), 'FFF2CC'),
        (range(9, 13), 'E2EFDA'),
        (range(13, 18), 'DDEBF7'),
        (range(18, 23), 'FCE4D6'),
        (range(23, 25), 'FFC000'),
    ]
    for col in range(1, 25):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        for col_range, color in color_groups:
            if col in col_range:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                break
    for row in range(2, worksheet.max_row + 1):
        for col in range(1, 25):
            cell = worksheet.cell(row=row, column=col)
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = thin_border
            for col_range, color in color_groups:
                if col in col_range:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    break
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max(8, min(max_length + 2, 50))
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_overview(coversheet_files: List[Path], output_path: str = 'overview_output.xlsx',
                   lookup_file_path: str = None):
    """
    Create overview Excel file from multiple coversheet files.
    A single xlwings App instance is created once and shared across all files
    to avoid repeated Excel COM process creation (which is slow and error-prone).
    """
    import xlwings as xw

    clin_lookup_table = {}
    if lookup_file_path and Path(lookup_file_path).exists():
        clin_lookup_table = load_clin_lookup_table(lookup_file_path)
    
    overview_columns = [
        'PR#', 'Version', 'OpDiv', 'Ticket', '# of CLINs (Line items in current OP)',
        'OpDiv Approval', 'Services', 'SF30 Description', 'F&R Needed', '# of CLINs',
        'F&R Notes', 'Type of F&R', 'Liaison', 'SME', 'Type of Mod',
        'Coversheet Description', 'Funding', '12M+ CLIN', 'MSA Mod', 'Standard Notes',
        'New Req?', 'Section C Update Needed?', 'Tara Matthews Approval', 'Cristy Scott Approval'
    ]
    
    rows = []

    # Create a single shared xlwings App instance for all checkbox reads
    xlwings_app = xw.App(visible=False, add_book=False)
    try:
        for filepath in coversheet_files:
            try:
                print(f"Processing: {filepath.name}")
                row_data = process_coversheet(filepath, clin_lookup_table, xlwings_app=xlwings_app)
                rows.append(row_data)
            except Exception as e:
                print(f"ERROR processing {filepath.name}: {e}")
    finally:
        xlwings_app.quit()
    
    df = pd.DataFrame(rows, columns=overview_columns)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Overview', index=False)
        worksheet = writer.sheets['Overview']
        apply_styling(worksheet)
    
    print(f"\nSuccessfully processed {len(rows)} out of {len(coversheet_files)} files.")
    return df



def configure_runtime():
    global BASE_DIR, INPUT_DIR, OUTPUT_DIR, COVERSHEET_FILES_DIR
    global clin_table_file, overview_output_file

    print("Select the coversheet folder and related files for overview generation...")
    coversheets_dir = choose_directory(
        title="Select the coversheets folder",
        state_key="script02_coversheets_dir",
    )
    clin_lookup_path = choose_file(
        title="Select the CLIN lookup workbook",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xlsb *.xls")],
        state_key="script02_clin_lookup_file",
    )
    overview_path = choose_save_file(
        title="Choose where to save the overview workbook",
        default_name="overview_file.xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        state_key="script02_overview_output_file",
    )

    COVERSHEET_FILES_DIR = coversheets_dir
    OUTPUT_DIR = overview_path.parent
    INPUT_DIR = clin_lookup_path.parent
    BASE_DIR = coversheets_dir.parent
    clin_table_file = clin_lookup_path
    overview_output_file = overview_path
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def main():
    """Main function to process coversheet files."""
    configure_runtime()
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if COVERSHEET_FILES_DIR.exists():
        coversheet_files = list(COVERSHEET_FILES_DIR.glob("*.xlsx"))
        coversheet_files = [f for f in coversheet_files
                           if 'overview' not in f.name.lower()
                           and not f.name.startswith("~$")]

        if coversheet_files:
            if clin_table_file.exists():
                df = create_overview(coversheet_files, str(overview_output_file),
                                     str(clin_table_file))
            else:
                print("WARNING: CLIN lookup file not found. Services will be blank.")
                df = create_overview(coversheet_files, str(overview_output_file), None)

            print("\nProcess completed successfully!")
        else:
            print(f"No Excel files found in {COVERSHEET_FILES_DIR}")
    else:
        print(f"Directory not found: {COVERSHEET_FILES_DIR}")
        print("Please ensure the directory path is correct.")


if __name__ == "__main__":
    main()