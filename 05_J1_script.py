"""
J1 Current File Automation

This script reads from the Catalog sheet in the Build file and appends entries
to the appropriate sheets in the J1 file based on TO Period.

The selected previous J.1 workbook is copied/converted to an .xlsx output
workbook before modification. Binary .xlsb/.xls baselines cannot simply be
renamed to .xlsx; when one is selected, the script attempts a real Excel
conversion before using openpyxl to append rows.
"""

import pandas as pd
import shutil
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import warnings
from file_selection import (
    choose_existing_file, choose_save_file, read_excel_auto, EXCEL_FILETYPES,
)

warnings.filterwarnings('ignore')


# ─── Runtime-selected inputs ──────────────────────────────────────────────────
# Assigned in main() from CLI arguments or file picker dialogs.
build_file = None
j1_previous_file_input = None
j1_current_file = None


# ============================================================
# CALCULATE CURRENT OPT PD
# ============================================================

def get_current_opt_pd(verbose=True):
    """
    Calculate the current Option Period based on the current date.
    
    Option Period Schedule:
    - Opt Pd 5: 9/1/2025 - 8/31/2026
    - Opt Pd 6: 9/1/2026 - 8/31/2027
    - Opt Pd 7: 9/1/2027 - 8/31/2028
    - Opt Pd 8: 9/1/2028 - 8/31/2029
    - Opt Pd 9: 9/1/2029 - 8/31/2030
    - Opt Pd 10: 9/1/2030 - 8/31/2031
    - Opt Pd 11: 9/1/2031 - 8/31/2032
    
    Returns:
        Current Option Period number (5-11)
    """
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month
    
    # Base year for Opt Pd 5 (starts September 1, 2025)
    base_year = 2025
    base_opt_pd = 5
    
    # Determine which period we're in
    # If we're in September or later, we're in the period starting that year
    # If we're before September, we're in the period that started last year
    if current_month >= 9:
        # We're in Sep-Dec, use current year as start
        years_since_base = current_year - base_year
    else:
        # We're in Jan-Aug, use previous year as start
        years_since_base = current_year - base_year - 1
    
    current_opt_pd = base_opt_pd + years_since_base
    
    # Format dates for display
    if current_month >= 9:
        period_start = f"9/1/{current_year}"
        period_end = f"8/31/{current_year + 1}"
    else:
        period_start = f"9/1/{current_year - 1}"
        period_end = f"8/31/{current_year}"
    
    if verbose:
        print(f"\nCurrent Date: {current_date.strftime('%m/%d/%Y %H:%M:%S')}")
        print(f"Current Option Period: {current_opt_pd}")
        print(f"Period Range: {period_start} - {period_end}")
    
    return current_opt_pd


# Store the current Opt Pd as a global variable
CURRENT_OPT_PD = get_current_opt_pd(verbose=False)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def create_lookup_value(row, include_to_period=False):
    """
    Create lookup value by concatenating relevant columns based on pricing method.
    Skips empty columns and preserves integer formatting.
    
    Args:
        row: DataFrame row
        include_to_period: If True, includes TO Period in the lookup value (for Lookup Value #2)
        
    Returns:
        Concatenated lookup value string
    """
    # Define lookup value logic for each pricing method
    if include_to_period:
        pricing_method_lookup_map = {
            'ICB': ['EIS CLIN', 'Case Number', 'SRE Pricing Element', 'TO Period'],
            'ORIG NSC': ['EIS CLIN', 'Orig NSC', 'SRE Pricing Element', 'TO Period'],
            'ORIG NSC - TERM NSC': ['EIS CLIN', 'Orig NSC', 'Term NSC', 'SRE Pricing Element', 'TO Period'],
            'TERM NSC': ['EIS CLIN', 'Term NSC', 'SRE Pricing Element', 'TO Period'],
            'ORIG JUR': ['EIS CLIN', 'Orig CJID', 'SRE Pricing Element', 'TO Period'],
            'ORIG JUR - TERM JUR': ['EIS CLIN', 'Orig CJID', 'Term CJID', 'SRE Pricing Element', 'TO Period'],
            'TERM JUR': ['EIS CLIN', 'Term CJID', 'SRE Pricing Element', 'TO Period'],
        }
    else:
        pricing_method_lookup_map = {
            'ICB': ['EIS CLIN', 'Case Number', 'SRE Pricing Element'],
            'ORIG NSC': ['EIS CLIN', 'Orig NSC', 'SRE Pricing Element'],
            'ORIG NSC - TERM NSC': ['EIS CLIN', 'Orig NSC', 'Term NSC', 'SRE Pricing Element'],
            'TERM NSC': ['EIS CLIN', 'Term NSC', 'SRE Pricing Element'],
            'ORIG JUR': ['EIS CLIN', 'Orig CJID', 'SRE Pricing Element'],
            'ORIG JUR - TERM JUR': ['EIS CLIN', 'Orig CJID', 'Term CJID', 'SRE Pricing Element'],
            'TERM JUR': ['EIS CLIN', 'Term CJID', 'SRE Pricing Element'],
        }
    
    pricing_method = str(row['YY_Price_Method']).strip()
    
    # Get the columns to concatenate for this pricing method
    columns_to_use = pricing_method_lookup_map.get(pricing_method, ['EIS CLIN'])
    
    # Concatenate values with "/" separator, skipping empty values
    lookup_parts = []
    for col in columns_to_use:
        value = row[col]
        
        # Skip if value is NaN or empty
        if pd.isna(value) or value == "":
            continue
        
        # Convert value to string while preserving integers
        if isinstance(value, float):
            # Check if it's actually an integer (e.g., 123495.0 -> 123495)
            if value.is_integer():
                value_str = str(int(value))
            else:
                value_str = str(value)
        else:
            value_str = str(value).strip()
        
        # Only add non-empty values
        if value_str:
            lookup_parts.append(value_str)
    
    return "/".join(lookup_parts)


def get_max_price_id(j1_file_path):
    """
    Find the maximum Price Id across all three sheets in the J1_Example file.
    
    Args:
        j1_file_path: Path to J1_Example file
        
    Returns:
        Maximum Price Id value
    """
    sheet_names = ['2A_Opt Pd 1-4 Catalog', '2B_Opt Pd 5 Catalog', '2C_Opt Pd 6-11 Catalog']
    max_price_id = 0
    
    for sheet_name in sheet_names:
        try:
            df = read_excel_auto(j1_file_path, sheet_name=sheet_name)
            if 'Price Id' in df.columns:
                sheet_max = df['Price Id'].max()
                if pd.notna(sheet_max) and sheet_max > max_price_id:
                    max_price_id = int(sheet_max)
        except Exception as e:
            print(f"Warning: Could not read {sheet_name}: {e}")
    
    print(f"Maximum Price Id found: {max_price_id}")
    return max_price_id


def read_catalog_from_build(build_file_path):
    """
    Read the Catalog sheet from the Build file.
    
    Args:
        build_file_path: Path to the Build file
        
    Returns:
        DataFrame with Catalog data
    """
    try:
        catalog_df = read_excel_auto(build_file_path, sheet_name='Catalog')
        catalog_df.columns = catalog_df.columns.str.strip()
        print(f"Read {len(catalog_df)} rows from Catalog sheet")
        return catalog_df
    except Exception as e:
        print(f"Error reading Catalog sheet: {e}")
        return pd.DataFrame()


def format_date_value(date_value):
    """
    Format date value to DD-MM-YYYY format.
    
    Args:
        date_value: Date value in various formats
        
    Returns:
        Formatted date string in DD-MM-YYYY format or original value
    """
    if pd.isna(date_value) or date_value == "":
        return ""
    
    try:
        if isinstance(date_value, str):
            date_obj = pd.to_datetime(date_value)
            return date_obj.strftime('%d-%m-%Y')
        
        if isinstance(date_value, pd.Timestamp):
            return date_value.strftime('%d-%m-%Y')
        
        date_obj = pd.to_datetime(date_value)
        return date_obj.strftime('%d-%m-%Y')
        
    except:
        return str(date_value)


def format_hhs_price(price_value):
    """
    Format HHS Price as a 6 decimal point float.
    
    Args:
        price_value: Price value (can be float, int, or string)
        
    Returns:
        Float value with 6 decimal places
    """
    if pd.isna(price_value) or price_value == "":
        return 0.000000
    
    try:
        # Convert to float and format with 6 decimal places
        price_float = float(price_value)
        return float(f"{price_float:.6f}")
    except (ValueError, TypeError):
        # If conversion fails, return 0
        return 0.000000


def prepare_j1_entry(catalog_row, price_id):
    """
    Prepare a single entry for the J1_Example file from a Catalog row.
    
    Args:
        catalog_row: Row from Catalog DataFrame
        price_id: Price Id to assign
        
    Returns:
        Dictionary with J1_Example columns
    """
    # Map Catalog columns to J1_Example columns
    j1_entry = {
        'YY_Price_Method': catalog_row.get('YY_Price_Method', ''),
        'EIS CLIN': catalog_row.get('EIS CLIN', ''),
        'EIS CLIN Name': catalog_row.get('EIS CLIN Name', ''),
        'Frequency': catalog_row.get('EIS Frequency', ''),  # Note: EIS Frequency -> Frequency
        'Charging Unit': catalog_row.get('Charging Unit', ''),
        'Orig NSC': catalog_row.get('Orig NSC', ''),
        'Term NSC': catalog_row.get('Term NSC', ''),
        'Orig CJID': catalog_row.get('Orig CJID', ''),
        'Term CJID': catalog_row.get('Term CJID', ''),
        'HHS Price': format_hhs_price(catalog_row.get('HHS Price', 0)),  # Format as 6 decimal float
        'wdm ICB': catalog_row.get('wdm ICB', ''),
        'wdm NSP': catalog_row.get('wdm NSP', ''),
        'Case Number': catalog_row.get('Case Number', ''),
        'Verizon Case Description': catalog_row.get('Verizon Case Description', ''),
        'SRE Pricing Element': catalog_row.get('SRE Pricing Element', ''),
        'SRE Device Class ID': catalog_row.get('SRE Device Class ID', ''),
        'HHS Discount from OLP': catalog_row.get('HHS Discount from OLP', ''),
        'Band Low': catalog_row.get('Band Low', ''),
        'Band High': catalog_row.get('Band High', ''),
        'TO Period': catalog_row.get('TO Period', ''),
        'Price Start Date': format_date_value(catalog_row.get('Price Start Date', '')),  # Format date
        'Price End Date': format_date_value(catalog_row.get('Price End Date', '')),  # Format date
        'Most Recent Mod': catalog_row.get('Mod Number', ''),  # Note: Mod Number -> Most Recent Mod
        'Most Recent Mod Notes': catalog_row.get('Price Request Number', ''),  # Note: Price Request Number -> Most Recent Mod Notes
        'Lookup Value': '',  # Will be calculated
        'Lookup Value #2': '',  # Will be calculated
        'Price Id': price_id
    }
    
    # Calculate lookup values
    j1_entry['Lookup Value'] = create_lookup_value(catalog_row, include_to_period=False)
    j1_entry['Lookup Value #2'] = create_lookup_value(catalog_row, include_to_period=True)
    
    return j1_entry


def categorize_by_to_period(catalog_df):
    """
    Categorize catalog entries by TO Period.
    
    Args:
        catalog_df: DataFrame with Catalog data
        
    Returns:
        Dictionary with keys 'pd5' and 'pd6_11' containing filtered DataFrames
    """
    # Convert TO Period to numeric for comparison
    catalog_df['TO_Period_Numeric'] = pd.to_numeric(catalog_df['TO Period'], errors='coerce')
    
    # Filter by the selected/current Option Period instead of hard-coded OP 5.
    current_entries = catalog_df[catalog_df['TO_Period_Numeric'] == CURRENT_OPT_PD].copy()
    future_entries = catalog_df[
        (catalog_df['TO_Period_Numeric'] >= CURRENT_OPT_PD + 1) &
        (catalog_df['TO_Period_Numeric'] <= 11)
    ].copy()
    
    print(f"\nCategorized entries:")
    print(f"  TO Period {CURRENT_OPT_PD}: {len(current_entries)} entries")
    print(f"  TO Period {CURRENT_OPT_PD + 1}-11: {len(future_entries)} entries")
    
    return {
        'pd5': current_entries,
        'pd6_11': future_entries
    }


def apply_yellow_formatting(worksheet, start_row, end_row, num_cols, hhs_price_col_index=None):
    """
    Apply yellow background and borders to newly added rows with left alignment.
    Also applies number formatting to HHS Price column.
    
    Args:
        worksheet: openpyxl worksheet
        start_row: First row to format
        end_row: Last row to format
        num_cols: Number of columns
        hhs_price_col_index: Column index for HHS Price (1-based)
    """
    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    left_alignment = Alignment(horizontal="left", vertical="center")  # Changed to left alignment
    
    # Apply formatting to each cell in the new rows
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = yellow_fill
            cell.border = thin_border
            cell.alignment = left_alignment  # Left align all values
            
            # Apply number format to HHS Price column
            if hhs_price_col_index and col == hhs_price_col_index:
                cell.number_format = '0.000000'  # Force 6 decimal places


def append_to_j1_sheet(j1_file_path, sheet_name, entries_df, start_price_id):
    """
    Append entries to a specific sheet in the J1_Example file with yellow highlighting.
    
    Args:
        j1_file_path: Path to J1_Example file
        sheet_name: Name of the sheet to append to
        entries_df: DataFrame with entries to append
        start_price_id: Starting Price Id for these entries
        
    Returns:
        Next available Price Id
    """
    if entries_df.empty:
        print(f"  No entries to append to {sheet_name}")
        return start_price_id
    
    print(f"\nAppending {len(entries_df)} entries to {sheet_name}...")
    
    # Load workbook
    wb = load_workbook(j1_file_path)
    ws = wb[sheet_name]
    
    # Find the last row with data
    last_row = ws.max_row
    
    # Get column order from the sheet header and find HHS Price column
    j1_columns = []
    hhs_price_col_index = None
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            header_str = str(header_value).strip()
            j1_columns.append(header_str)
            if header_str == 'HHS Price':
                hhs_price_col_index = col
    
    print(f"  Starting at row {last_row + 1}")
    print(f"  Price Id range: {start_price_id} to {start_price_id + len(entries_df) - 1}")
    
    # Prepare entries with Price Ids
    current_price_id = start_price_id
    new_rows_start = last_row + 1
    
    for idx, catalog_row in entries_df.iterrows():
        # Prepare J1 entry
        j1_entry = prepare_j1_entry(catalog_row, current_price_id)
        
        # Write to worksheet
        row_num = last_row + 1
        for col_idx, col_name in enumerate(j1_columns, start=1):
            value = j1_entry.get(col_name, '')
            ws.cell(row=row_num, column=col_idx, value=value)
        
        last_row += 1
        current_price_id += 1
    
    # Apply yellow formatting and borders to new rows, including HHS Price number format
    apply_yellow_formatting(ws, new_rows_start, last_row, len(j1_columns), hhs_price_col_index)
    
    # Save workbook
    wb.save(j1_file_path)
    wb.close()
    
    print(f"Successfully appended {len(entries_df)} entries")
    
    return current_price_id



def _force_xlsx_output_path(path):
    """Ensure the J.1 output path is an .xlsx workbook path."""
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        corrected = path.with_suffix(".xlsx")
        print(f"Warning: J.1 output must be .xlsx for automated updates; using {corrected}")
        return corrected
    return path


def _convert_with_excel_com(source_path, output_path):
    """
    Convert .xlsb/.xls source workbook to .xlsx using desktop Excel.

    openpyxl cannot read or write .xlsb workbooks. On Windows workstations with
    Excel installed, pywin32 can drive Excel's own SaveAs conversion so the
    existing workbook structure is preserved before rows are appended.
    """
    source_path = Path(source_path)
    output_path = Path(output_path)

    try:
        import win32com.client as win32
    except Exception as exc:
        raise RuntimeError(
            "The selected previous J.1 workbook is .xlsb/.xls, which must be "
            "converted to .xlsx before the automation can append rows. Install "
            "pywin32 and run this step on a Windows machine with Microsoft Excel, "
            "or manually open the previous J.1 workbook in Excel, Save As .xlsx, "
            "and rerun Step 5 using that .xlsx file."
        ) from exc

    excel = None
    workbook = None
    try:
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.DisplayAlerts = False
        excel.Visible = False

        output_path.parent.mkdir(parents=True, exist_ok=True)
        if output_path.exists():
            output_path.unlink()

        workbook = excel.Workbooks.Open(str(source_path.resolve()), ReadOnly=True)
        # Excel FileFormat 51 = xlOpenXMLWorkbook (.xlsx)
        workbook.SaveAs(str(output_path.resolve()), FileFormat=51)
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.DisplayAlerts = True
            excel.Quit()

    if not output_path.exists():
        raise RuntimeError(f"Excel conversion did not create expected output: {output_path}")


def prepare_j1_current_workbook(previous_path, output_path):
    """
    Create an editable .xlsx current J.1 workbook from the selected prior J.1.

    .xlsx/.xlsm inputs can be copied directly because openpyxl can read the
    workbook package. .xlsb/.xls inputs require a real conversion instead of a
    byte-for-byte copy because renaming an .xlsb file to .xlsx creates the
    'File contains no valid workbook part' error seen in Step 5.
    """
    previous_path = Path(previous_path)
    output_path = _force_xlsx_output_path(output_path)

    if not previous_path.exists():
        raise FileNotFoundError(f"J1 previous file not found at {previous_path}")

    suffix = previous_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        shutil.copy2(previous_path, output_path)
        print(f"Copied J1 previous file to: {output_path}")
    elif suffix in {".xlsb", ".xls"}:
        print(f"Converting J1 previous file from {suffix} to .xlsx: {output_path}")
        _convert_with_excel_com(previous_path, output_path)
        print(f"Converted J1 previous file to: {output_path}")
    else:
        raise ValueError(f"Unsupported J.1 previous file type: {previous_path.suffix}")

    # Validate the output before later steps attempt to append rows.
    try:
        wb = load_workbook(output_path, read_only=True)
        wb.close()
    except Exception as exc:
        raise RuntimeError(
            f"The prepared J.1 output workbook is not a valid .xlsx file: {output_path}. "
            "If the source file was .xlsb/.xls, confirm it was converted rather than renamed."
        ) from exc

    return output_path

# ============================================================
# MAIN EXECUTION
# ============================================================

def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Create the current J.1 workbook from selected prior J.1 and build workbooks."
    )
    parser.add_argument("--build-file", help="Build workbook containing the Catalog sheet")
    parser.add_argument("--j1-previous-file", help="Prior/current baseline J.1 workbook to copy and update")
    parser.add_argument("--output-file", help="Output J.1 workbook path")
    parser.add_argument("--current-opt-pd", type=int, default=CURRENT_OPT_PD,
                        help="Current option period; defaults to calculated value")
    return parser.parse_args(argv)


def configure_paths(args):
    global build_file, j1_previous_file_input, j1_current_file, CURRENT_OPT_PD
    CURRENT_OPT_PD = args.current_opt_pd
    build_file = choose_existing_file(args.build_file, "Select the build workbook with the Catalog sheet", EXCEL_FILETYPES)
    j1_previous_file_input = choose_existing_file(args.j1_previous_file, "Select the J.1 previous workbook", EXCEL_FILETYPES)
    j1_current_file = choose_save_file(args.output_file, "Save the updated/current J.1 workbook as", "j1_current_file.xlsx", EXCEL_FILETYPES)
    j1_current_file = _force_xlsx_output_path(j1_current_file)


def main(argv=None):
    """Main function to execute the J1 automation."""
    global j1_current_file

    args = parse_args(argv)
    configure_paths(args)
    j1_current_file.parent.mkdir(parents=True, exist_ok=True)

    # Copy or convert the selected J.1 previous file before modifying.
    try:
        j1_current_file = prepare_j1_current_workbook(j1_previous_file_input, j1_current_file)
    except Exception as exc:
        print(f"Error preparing J1 current workbook: {exc}")
        return

    print("=" * 60)
    print("J1 PREVIOUS FILE AUTOMATION")
    print("=" * 60)
    print(f"Build File (Catalog source): {build_file}")
    print(f"J1 Current File: {j1_current_file}")
    print(f"\nCurrent Option Period: {CURRENT_OPT_PD}")
    print("=" * 60)

    # Step 1: Read Catalog from Build file
    print("\nStep 1: Reading Catalog sheet from Build file...")
    catalog_df = read_catalog_from_build(build_file)

    if catalog_df.empty:
        print("Error: No data found in Catalog sheet")
        return

    # Step 2: Get maximum Price Id
    print("\nStep 2: Finding maximum Price Id...")
    max_price_id = get_max_price_id(j1_current_file)
    next_price_id = max_price_id + 1

    # Step 3: Categorize entries by TO Period
    print("\nStep 3: Categorizing entries by TO Period...")
    categorized = categorize_by_to_period(catalog_df)

    # Step 4: Append to current OP sheet
    print("\nStep 4: Appending entries to sheets...")
    next_price_id = append_to_j1_sheet(
        j1_current_file,
        f'2B_Opt Pd {CURRENT_OPT_PD} Catalog',
        categorized['pd5'],
        next_price_id
    )

    # Step 5: Append to future OP sheet
    next_price_id = append_to_j1_sheet(
        j1_current_file,
        f'2C_Opt Pd {CURRENT_OPT_PD + 1}-11 Catalog',
        categorized['pd6_11'],
        next_price_id
    )

    # Final summary
    print("\n" + "=" * 60)
    print("AUTOMATION COMPLETE")
    print("=" * 60)
    print(f"Total entries processed: {len(catalog_df)}")
    print(f"  - TO Period {CURRENT_OPT_PD}: {len(categorized['pd5'])} entries")
    print(f"  - TO Period {CURRENT_OPT_PD + 1}-11: {len(categorized['pd6_11'])} entries")
    print(f"Final Price Id: {next_price_id - 1}")
    print(f"J1 current file updated: {j1_current_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()